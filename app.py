import os
import json
import random
from datetime import datetime
from flask import Flask, render_template, redirect, url_for, request, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from sqlalchemy import inspect, text, event
from sqlalchemy.engine import Engine
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from models import (
    db, User, Department, Faculty, Student, Admin,
    PreAdvisingCourse, SectionOffering, AdvisingWindow, AdvisingPlan, Registration,
    SemesterDropRequest, AttendanceRecord, AdvisingRequest, Grade, LedgerEntry, Installment, Announcement, SystemSetting
)

@event.listens_for(Engine, "connect")
def set_sqlite_pragma(dbapi_connection, connection_record):
    import sqlite3
    if isinstance(dbapi_connection, sqlite3.Connection):
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.execute("PRAGMA synchronous=NORMAL")
        cursor.close()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ewu_secret_key_123')

# Database URI — defaults to SQLite (zero setup). Set DATABASE_URL env var for PostgreSQL.
DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///ewu_portal.db')
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# Prevent disk I/O errors: disable connection pool for SQLite (each request gets its own connection)
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {'check_same_thread': False, 'timeout': 30},
    'pool_recycle': 300,
    'pool_pre_ping': True,
}

db.init_app(app)

PROFILE_UPLOAD_DIR = os.path.join(app.root_path, 'static', 'uploads')
ALLOWED_PROFILE_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def ensure_runtime_schema():
    """Create tables and add lightweight columns needed by existing installs."""
    db.create_all()
    inspector = inspect(db.engine)
    tables = set(inspector.get_table_names())
    profile_columns = {
        'students': {
            'profile_pic': 'VARCHAR(255)', 
            'about': 'VARCHAR(500)',
            'phone_number': 'VARCHAR(50)',
            'remaining_credits': 'FLOAT DEFAULT 0.0',
            'present_address': 'VARCHAR(255)',
            'permanent_address': 'VARCHAR(255)',
            'completed_courses_and_grades': 'TEXT',
            'current_courses': 'TEXT',
            'current_course_credit': 'FLOAT DEFAULT 0.0',
            'next_semester_courses': 'TEXT',
            'next_semester_course_credit': 'FLOAT DEFAULT 0.0'
        },
        'faculty': {'profile_pic': 'VARCHAR(255)', 'about': 'VARCHAR(500)'},
        'pre_advising_courses': {'completed_credit_requirement': 'INTEGER DEFAULT 0'},
        'section_offerings': {'completed_credit_requirement': 'INTEGER DEFAULT 0'},
    }

    for table_name, columns in profile_columns.items():
        if table_name not in tables:
            continue
        existing = {column['name'] for column in inspector.get_columns(table_name)}
        for column_name, column_type in columns.items():
            if column_name not in existing:
                db.session.execute(text(f'ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}'))
    db.session.commit()

with app.app_context():
    ensure_runtime_schema()

# Register custom Jinja2 filters
@app.template_filter('max_val')
def max_val_filter(value, other):
    return max(value, other)

app.jinja_env.filters['max_val'] = max_val_filter

login_manager = LoginManager()
login_manager.login_view = 'login_page'
login_manager.login_message = 'Please sign in to continue.'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    u = User.query.get(user_id)
    if u and u.is_active:
        return u
    return None

# Check deactivated users on every request
@app.before_request
def check_deactivated_user():
    if request.path.startswith('/static'):
        return
    if current_user.is_authenticated:
        u = User.query.get(current_user.id)
        if not u or not u.is_active:
            logout_user()
            flash('Your account has been deactivated by the administrator.', 'error')
            return redirect(url_for('login_page'))

# Helper: Get current system semester
def get_current_semester():
    sem_setting = SystemSetting.query.filter_by(key='current_semester').first()
    return sem_setting.value if sem_setting else 'Spring2026'

# Helper: Get next system semester for advising registration
def get_next_semester():
    sem_setting = SystemSetting.query.filter_by(key='next_semester').first()
    return sem_setting.value if sem_setting else 'Summer2026'

# Helper: Get calendar dates
def get_calendar_dates():
    c_start = SystemSetting.query.filter_by(key='current_semester_start').first()
    c_end = SystemSetting.query.filter_by(key='current_semester_end').first()
    n_start = SystemSetting.query.filter_by(key='next_semester_start').first()
    n_end = SystemSetting.query.filter_by(key='next_semester_end').first()
    return {
        'current_semester_start_date': c_start.value if c_start else '2026-01-05',
        'current_semester_end_date': c_end.value if c_end else '2026-04-20',
        'next_semester_start_date': n_start.value if n_start else '2026-05-10',
        'next_semester_end_date': n_end.value if n_end else '2026-08-25',
    }

# Helper: Get prev semester for schedules comparison
def get_previous_semester():
    curr = get_current_semester()
    if curr == 'Summer2026': return 'Spring2026'
    if curr == 'Spring2026': return 'Fall2025'
    return 'Summer2026'

def save_profile_pic_upload(file_storage, owner_prefix):
    if not file_storage or not file_storage.filename:
        return None

    raw_filename = secure_filename(file_storage.filename)
    if '.' not in raw_filename:
        raise ValueError('Invalid image format. Allowed: png, jpg, jpeg, gif, webp.')

    stem, extension = os.path.splitext(raw_filename)
    extension = extension.lstrip('.').lower()
    if extension not in ALLOWED_PROFILE_IMAGE_EXTENSIONS:
        raise ValueError('Invalid image format. Allowed: png, jpg, jpeg, gif, webp.')

    safe_owner = secure_filename(owner_prefix) or 'profile'
    safe_stem = stem or 'photo'
    timestamp = datetime.utcnow().strftime('%Y%m%d%H%M%S')
    unique_filename = f"{safe_owner}_{timestamp}_{random.randint(1000, 9999)}_{safe_stem}.{extension}"
    os.makedirs(PROFILE_UPLOAD_DIR, exist_ok=True)
    file_storage.save(os.path.join(PROFILE_UPLOAD_DIR, unique_filename))
    return unique_filename

MAJOR_CREDITS = {
    'CSE': 140.0,
    'EEE': 140.0,
    'ICE': 140.0,
    'CEN': 140.0, # Civil Engineering
    'GEB': 140.0, # Genetic Engineering & Biotech
    'PHR': 170.0, # Pharmacy
    'DSA': 130.0, # Data Science & Analytics
    'MAT': 130.0, # Mathematics
    'BBA': 130.0,
    'ECO': 130.0,
    'ENG': 130.0,
    'LAW': 130.0,
    'SOC': 130.0,
    'INF': 130.0,
    'PPHS': 130.0
}

GRADE_POINTS_MAP = {
    'A': 4.0, 'A-': 3.7, 'B+': 3.3, 'B': 3.0, 'B-': 2.7,
    'C+': 2.3, 'C': 2.0, 'C-': 1.7, 'D+': 1.3, 'D': 1.0, 'F': 0.0
}

DAY_NAMES = {
    'S': 'Saturday',
    'M': 'Monday',
    'T': 'Tuesday',
    'W': 'Wednesday',
    'R': 'Thursday',
    'F': 'Friday',
    'U': 'Sunday',
}

DAY_ORDER = {
    'Saturday': 0,
    'Sunday': 1,
    'Monday': 2,
    'Tuesday': 3,
    'Wednesday': 4,
    'Thursday': 5,
    'Friday': 6,
}

# Helper: parse schedule conflicts
def parse_schedule(schedule_str):
    try:
        if not schedule_str or ':' not in schedule_str:
            return []
        # Split by the FIRST colon only
        days_part, time_part = schedule_str.split(':', 1)
        days_part = days_part.strip()
        time_part = time_part.strip()
        
        days = []
        i = 0
        while i < len(days_part):
            days.append(days_part[i])
            i += 1
                
        start_str, end_str = time_part.split('-')
        
        def to_minutes(time_str):
            # Normalizes times like '10.10' or '10:10' or '10.10.1' (admin typo support)
            clean = time_str.strip().replace(':', '.')
            parts = clean.split('.')
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            return h * 60 + m
            
        start_min = to_minutes(start_str)
        end_min = to_minutes(end_str)
        
        return [(d, start_min, end_min) for d in days]
    except Exception:
        return []

def schedules_conflict(s1, s2):
    p1 = parse_schedule(s1)
    p2 = parse_schedule(s2)
    for d1, start1, end1 in p1:
        for d2, start2, end2 in p2:
            if d1 == d2:
                if start1 < end2 and start2 < end1:
                    return True
    return False

def format_minutes(total_minutes):
    hour = total_minutes // 60
    minute = total_minutes % 60
    suffix = 'AM' if hour < 12 else 'PM'
    display_hour = hour % 12 or 12
    return f"{display_hour}:{minute:02d} {suffix}"

def expand_section_schedule(section):
    items = []
    for day_code, start_min, end_min in parse_schedule(section.schedule):
        day_name = DAY_NAMES.get(day_code, day_code)
        items.append({
            'day': day_name,
            'day_sort': DAY_ORDER.get(day_name, 99),
            'start_sort': start_min,
            'time': f"{format_minutes(start_min)} - {format_minutes(end_min)}",
            'section': section,
        })
    return items

def section_to_dict(section):
    return {
        'id': section.id,
        'course_code': section.course_code,
        'course_title': section.course_title,
        'section_number': section.section_number,
        'credits': section.credits,
        'schedule': section.schedule,
        'room': section.room,
        'capacity': section.capacity,
        'enrolled_count': section.enrolled_count,
        'is_lab': section.is_lab,
        'linked_section_id': section.linked_section_id,
        'faculty_id': section.faculty_id,
    }

def current_faculty_profile():
    if not current_user.is_authenticated or current_user.role != 'faculty':
        return None
    return Faculty.query.filter_by(user_id=current_user.id).first()

def faculty_can_teach_section(faculty, section):
    return bool(faculty and section and section.faculty_id == faculty.id)

def registered_students_for_section(section_id):
    registrations = Registration.query.filter_by(
        section_id=section_id
    ).all()
    students = []
    for reg in registrations:
        student = Student.query.get(reg.student_id)
        if student:
            students.append(student)
    return sorted(students, key=lambda item: item.id)

def find_requested_section(req):
    if not req:
        return None

    section = SectionOffering.query.get(req.section_id) if req.section_id else None
    if section:
        return section

    if req.section_id:
        normalized_section = str(req.section_id).strip()
        section = SectionOffering.query.filter_by(
            course_code=req.course_id,
            section_number=normalized_section,
            semester_id=req.semester_id
        ).first()
        if section:
            return section

    return SectionOffering.query.filter_by(
        course_code=req.course_id,
        semester_id=req.semester_id
    ).order_by(SectionOffering.enrolled_count.asc()).first()

def student_has_schedule_conflict(student_id, candidate_section, exclude_section_id=None):
    registrations = Registration.query.filter_by(
        student_id=student_id,
        semester_id=candidate_section.semester_id
    ).all()
    exclude_ids = []
    if exclude_section_id:
        if isinstance(exclude_section_id, (list, tuple, set)):
            exclude_ids = list(exclude_section_id)
        else:
            exclude_ids = [exclude_section_id]
            
    for reg in registrations:
        if reg.section_id in exclude_ids:
            continue
        current_section = SectionOffering.query.get(reg.section_id)
        if current_section and schedules_conflict(current_section.schedule, candidate_section.schedule):
            return current_section
    return None

def save_grade_for_student(student_id, course_code, grade_letter, semester_id):
    grade_letter = (grade_letter or '').strip().upper()
    if grade_letter not in GRADE_POINTS_MAP:
        return False

    grade = Grade.query.filter_by(
        student_id=student_id,
        section_id=course_code,
        semester_id=semester_id
    ).first()
    if not grade:
        grade = Grade(
            id=f"grade-{student_id}-{course_code}-{semester_id}",
            student_id=student_id,
            section_id=course_code,
            grade_letter=grade_letter,
            grade_point=GRADE_POINTS_MAP[grade_letter],
            semester_id=semester_id
        )
        db.session.add(grade)
    else:
        grade.grade_letter = grade_letter
        grade.grade_point = GRADE_POINTS_MAP[grade_letter]
    return True

# Helper: check student credit-bracket gating
def is_student_allowed_in_portal(student):
    now_str = datetime.now().strftime('%Y-%m-%dT%H:%M')
    semester = get_current_semester()
    
    all_windows = AdvisingWindow.query.filter_by(semester_id=semester).all()
    if not all_windows:
        return False, None
        
    student_windows = [w for w in all_windows if w.credit_min <= student.completed_credits <= w.credit_max]
    if not student_windows:
        return False, None
        
    for w in student_windows:
        if w.start_date_time <= now_str <= w.end_date_time:
            return True, w
            
    return False, None

def get_active_window(student_credits, win_type):
    now_str = datetime.now().strftime('%Y-%m-%dT%H:%M')
    semester = get_current_semester()
    windows = AdvisingWindow.query.filter_by(type=win_type, semester_id=semester).all()
    for w in windows:
        if w.credit_min <= student_credits <= w.credit_max:
            if w.start_date_time <= now_str <= w.end_date_time:
                return w
    return None

# Context Processor
@app.context_processor
def inject_layout_variables():
    if not current_user.is_authenticated:
        return {}
    
    role_colors = {
        'student': '#3b82f6',
        'faculty': '#8b5cf6',
        'admin': '#f59e0b',
    }
    role_icons = {
        'student': 'graduation-cap',
        'faculty': 'users',
        'admin': 'settings',
    }
    
    role = current_user.role
    profile_name = "User"
    profile_id = current_user.id
    profile_pic = None
    profile_about = ""
    
    if role == 'student':
        student = Student.query.filter_by(user_id=current_user.id).first()
        if student:
            profile_name = student.name
            profile_id = student.id
            profile_pic = student.profile_pic
            profile_about = student.about or ""
    elif role == 'faculty':
        faculty = Faculty.query.filter_by(user_id=current_user.id).first()
        if faculty:
            profile_name = faculty.name
            profile_id = faculty.id
            profile_pic = faculty.profile_pic
            profile_about = faculty.about or ""
    elif role == 'admin':
        admin = Admin.query.filter_by(user_id=current_user.id).first()
        if admin:
            profile_name = admin.name
            profile_id = admin.id
            
    cal_dates = get_calendar_dates()
    return {
        'role_color': role_colors.get(role, '#3b82f6'),
        'role_icon': role_icons.get(role, 'user'),
        'profile_name': profile_name,
        'profile_id': profile_id,
        'profile_pic': profile_pic,
        'profile_about': profile_about,
        'current_semester': get_current_semester(),
        'next_semester': get_next_semester(),
        'current_semester_start_date': cal_dates['current_semester_start_date'],
        'current_semester_end_date': cal_dates['current_semester_end_date'],
        'next_semester_start_date': cal_dates['next_semester_start_date'],
        'next_semester_end_date': cal_dates['next_semester_end_date'],
    }

# ROUTES
@app.route('/')
def home():
    if current_user.is_authenticated:
        return redirect(url_for(current_user.role + '_dashboard'))
    return redirect(url_for('login_page'))

@app.route('/login', methods=['GET'])
def login_page():
    if current_user.is_authenticated:
        return redirect(url_for(current_user.role + '_dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def do_login():
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '')
    
    user = User.query.filter_by(email=email).first()
    if user:
        if not user.is_active:
            flash('Your account has been deactivated.', 'error')
            return redirect(url_for('login_page'))
        if not user.is_activated:
            flash('Account is not activated yet. Please click the Activation link.', 'error')
            return redirect(url_for('login_page'))
            
        if check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for(user.role + '_dashboard'))
            
    flash('Invalid credentials or deactivated account.', 'error')
    return redirect(url_for('login_page'))

# ACTIVATION FLOW
@app.route('/activate', methods=['GET', 'POST'])
def activate_account():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        user = User.query.filter_by(email=email).first()
        if not user:
            flash('Email not found in student/faculty records.', 'error')
            return redirect(url_for('activate_account'))
        
        # Generate 6 digit OTP
        code = str(random.randint(100000, 999999))
        user.otp_code = code
        db.session.commit()
        
        flash(f"Verification code successfully sent to email! [DEMO VERIFICATION CODE: {code}]", 'success')
        return render_template('activate_verify.html', email=email)
        
    return render_template('activate.html')

@app.route('/activate/verify', methods=['POST'])
def activate_verify():
    email = request.form.get('email', '').strip()
    code = request.form.get('code', '').strip()
    
    user = User.query.filter_by(email=email).first()
    if user and user.otp_code == code:
        return render_template('activate_password.html', email=email)
        
    flash('Incorrect verification code. Please try again.', 'error')
    return render_template('activate_verify.html', email=email)

@app.route('/activate/complete', methods=['POST'])
def activate_complete():
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '')
    confirm = request.form.get('confirm_password', '')
    
    if password != confirm:
        flash('Passwords do not match.', 'error')
        return render_template('activate_password.html', email=email)
        
    user = User.query.filter_by(email=email).first()
    if user:
        user.password_hash = generate_password_hash(password)
        user.is_activated = True
        user.otp_code = None
        db.session.commit()
        flash('Account activated successfully! You can now log in.', 'success')
        return redirect(url_for('login_page'))
        
    flash('User record error.', 'error')
    return redirect(url_for('login_page'))

# FORGOT PASSWORD FLOW
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        user = User.query.filter_by(email=email).first()
        if not user:
            flash('Email not found.', 'error')
            return redirect(url_for('forgot_password'))
            
        code = str(random.randint(100000, 999999))
        user.otp_code = code
        db.session.commit()
        flash(f"Reset code successfully generated! [DEMO RESET CODE: {code}]", 'success')
        return render_template('forgot_verify.html', email=email)
    return render_template('forgot.html')

@app.route('/forgot-password/verify', methods=['POST'])
def forgot_verify():
    email = request.form.get('email', '').strip()
    code = request.form.get('code', '').strip()
    
    user = User.query.filter_by(email=email).first()
    if user and user.otp_code == code:
        return render_template('forgot_password.html', email=email)
        
    flash('Incorrect reset code. Please try again.', 'error')
    return render_template('forgot_verify.html', email=email)

@app.route('/forgot-password/complete', methods=['POST'])
def forgot_complete():
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '')
    confirm = request.form.get('confirm_password', '')
    
    if password != confirm:
        flash('Passwords do not match.', 'error')
        return render_template('forgot_password.html', email=email)
        
    user = User.query.filter_by(email=email).first()
    if user:
        user.password_hash = generate_password_hash(password)
        user.otp_code = None
        db.session.commit()
        flash('Password successfully reset! You can now log in.', 'success')
        return redirect(url_for('login_page'))
        
    flash('User record error.', 'error')
    return redirect(url_for('login_page'))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login_page'))

# STUDENT DASHBOARD
@app.route('/student')
@login_required
def student_dashboard():
    return render_student_portal('dashboard')

@app.route('/advising')
@login_required
def student_advising():
    return render_student_portal('advising')

def get_eligible_courses_for_student(student):
    """Returns a set of course codes student is eligible to take."""
    # Build set of PASSED course codes (not section IDs)
    grades = Grade.query.filter_by(student_id=student.id).all()
    passed_codes = set()
    for g in grades:
        if g.grade_letter not in ('F', None, ''):
            # g.section_id might be a course_code directly stored, or a section ID
            # Try to look up the section to get course_code
            sec = SectionOffering.query.get(g.section_id)
            if sec:
                passed_codes.add(sec.course_code)
            else:
                # Fallback: treat section_id as the course_code (old format)
                passed_codes.add(g.section_id)

    # Current semester registrations (courses being taken now)
    current_regs = Registration.query.filter_by(
        student_id=student.id, semester_id=get_current_semester()
    ).all()
    current_course_codes = []
    ongoing_credits = 0.0
    for r in current_regs:
        sec_obj = SectionOffering.query.get(r.section_id)
        if sec_obj:
            current_course_codes.append(sec_obj.course_code)
            ongoing_credits += sec_obj.credits

    all_completed_or_ongoing = passed_codes | set(current_course_codes)
    student_total_credits = student.completed_credits + ongoing_credits
    
    # Define categories
    COMMON_COURSES = {
        'CHE109', 'CHE109 Lab', 'ENG101', 'ENG102', 'GEN226', 'MAT101', 
        'MAT102', 'MAT104', 'MAT205', 'PHY109', 'PHY109 Lab', 'PHY209', 'SAT101'
    }
    NON_DEPT_COURSES = {
        'ACT101', 'BUS231', 'BUS321', 'ECO101', 'FIN101', 'GEN210', 
        'MGT321', 'MGT337', 'MKT101'
    }
    
    # Count how many non-dept courses student has already completed or is currently taking
    completed_non_dept_count = sum(1 for c in all_completed_or_ongoing if c in NON_DEPT_COURSES)
    
    courses = PreAdvisingCourse.query.all()
    eligible_codes = set()
    
    for c in courses:
        # Rule 1: Department Check
        # If it's a common course or non-departmental course, it is allowed (with limits)
        is_common = c.code in COMMON_COURSES
        is_non_dept = c.code in NON_DEPT_COURSES
        
        # Core Department Check:
        # E.g. starts with CSE and student is CSE -> Core
        # If it starts with EEE/CSE/ICE and does not match student dept -> Other Core
        is_core_other = False
        if c.code.startswith('CSE') and student.department_id != 'CSE':
            is_core_other = True
        elif c.code.startswith('EEE') and student.department_id != 'EEE':
            is_core_other = True
        elif c.code.startswith('ICE') and student.department_id != 'ICE':
            is_core_other = True
            
        if is_core_other:
            continue
            
        # Non-Departmental Limit: at most 3 courses (9 credits)
        if is_non_dept and c.code not in all_completed_or_ongoing:
            if completed_non_dept_count >= 3:
                continue
                
        # Rule 2: Prerequisite Check
        prereqs_met = True
        for pre in c.prerequisites:
            if pre not in all_completed_or_ongoing:
                prereqs_met = False
                break
        if not prereqs_met:
            continue
            
        # Rule 3: Completed Credit Requirement Check
        if student_total_credits < c.completed_credit_requirement:
            continue
            
        eligible_codes.add(c.code)
        
    return eligible_codes

def render_student_portal(active_tab):
    if current_user.role != 'student':
        return redirect(url_for('home'))
        
    student = Student.query.filter_by(user_id=current_user.id).first()
    if not student:
        flash("Student profile not found. Please contact the administrator to create your record.", "error")
        logout_user()
        return redirect(url_for('login_page'))
    
    dept = Department.query.get(student.department_id)
    if not dept:
        class DummyDept:
            name = student.department_id or "Not Assigned"
        dept = DummyDept()
    advisor = Faculty.query.get(student.advisor_id) if student.advisor_id else None
    advisor_email = None
    if advisor:
        advisor_user = User.query.get(advisor.user_id)
        if advisor_user:
            advisor_email = advisor_user.email
    announcements = Announcement.query.filter(Announcement.target_role.in_(['student', 'all'])).all()
    
    # Calculate required & remaining credits
    total_required = MAJOR_CREDITS.get(student.department_id, 140.0)
    remaining_credits = max(0.0, total_required - student.completed_credits)
    
    # Pre-advising data (filtered by eligibility)
    eligible_course_codes = get_eligible_courses_for_student(student)
    raw_courses = [c for c in PreAdvisingCourse.query.all() if c.code in eligible_course_codes]
    all_catalog_map = {c.code: c for c in PreAdvisingCourse.query.all()}
    
    courses = []
    processed_codes = set()
    
    for c in raw_courses:
        if c.code in processed_codes:
            continue
            
        if " Lab" in c.code:
            theory_code = c.code.replace(" Lab", "")
            if theory_code in eligible_course_codes:
                continue
            else:
                courses.append(c)
                processed_codes.add(c.code)
        else:
            lab_code = c.code + " Lab"
            if lab_code in all_catalog_map:
                lab_course = all_catalog_map[lab_code]
                
                class CombinedCourse:
                    def __init__(self, code, title, credits, prerequisites, completed_credit_requirement):
                        self.code = code
                        self.title = title
                        self.credits = credits
                        self.prerequisites = prerequisites
                        self.completed_credit_requirement = completed_credit_requirement
                
                combined_prereqs = list(set(c.prerequisites) | set(lab_course.prerequisites))
                combined_credits = c.credits + lab_course.credits
                
                combined_obj = CombinedCourse(
                    code=c.code,
                    title=c.title + " & Lab",
                    credits=combined_credits,
                    prerequisites=combined_prereqs,
                    completed_credit_requirement=c.completed_credit_requirement
                )
                courses.append(combined_obj)
                processed_codes.add(c.code)
                processed_codes.add(lab_code)
            else:
                courses.append(c)
                processed_codes.add(c.code)
    
    all_courses = PreAdvisingCourse.query.all()
    all_courses_dict = {c.code: c for c in all_courses}
    
    courses_json = json.dumps([{
        'code': c.code,
        'title': c.title,
        'credits': c.credits,
        'prerequisites': c.prerequisites
    } for c in courses])
 
    plan = AdvisingPlan.query.filter_by(student_id=student.id, semester_id=get_next_semester()).first()
    plan_course_ids = plan.course_ids if plan else []
    plan_credits = sum([all_courses_dict[code].credits for code in plan_course_ids if code in all_courses_dict])
    
    # Offering sections
    registrations = Registration.query.filter_by(student_id=student.id, semester_id=get_next_semester()).all()
    selected_section_ids = [r.section_id for r in registrations]
    
    sections = [s for s in SectionOffering.query.filter_by(semester_id=get_next_semester()).all()
                if s.course_code in eligible_course_codes or s.id in selected_section_ids]
                
    sections_json = json.dumps([{
        'id': s.id,
        'course_code': s.course_code,
        'course_title': s.course_title,
        'section_number': s.section_number,
        'credits': s.credits,
        'schedule': s.schedule,
        'room': s.room,
        'dedicated_departments': s.dedicated_departments,
        'capacity': s.capacity,
        'enrolled_count': s.enrolled_count,
        'prerequisites': s.prerequisites,
        'is_lab': s.is_lab,
        'linked_section_id': s.linked_section_id,
        'faculty_id': s.faculty_id
    } for s in sections])
    
    # System Advising States (Toggle switches controlled by admin settings)
    # System Advising States (pre and final determined dynamically by window timings & credit requirements)
    now_str = datetime.now().strftime('%Y-%m-%dT%H:%M')
    
    pre_advising_active = False
    student_pre_window = AdvisingWindow.query.filter(
        AdvisingWindow.type == 'pre',
        AdvisingWindow.semester_id == get_current_semester(),
        AdvisingWindow.credit_min <= student.completed_credits,
        AdvisingWindow.credit_max >= student.completed_credits,
        AdvisingWindow.start_date_time <= now_str,
        AdvisingWindow.end_date_time >= now_str
    ).first()
    if student_pre_window:
        pre_advising_active = True
        
    final_advising_active = False
    student_final_window = AdvisingWindow.query.filter(
        AdvisingWindow.type == 'final',
        AdvisingWindow.semester_id == get_current_semester(),
        AdvisingWindow.credit_min <= student.completed_credits,
        AdvisingWindow.credit_max >= student.completed_credits,
        AdvisingWindow.start_date_time <= now_str,
        AdvisingWindow.end_date_time >= now_str
    ).first()
    if student_final_window:
        final_advising_active = True
        
    def is_setting_true(key):
        s = SystemSetting.query.filter_by(key=key).first()
        return s.value == 'true' if s else False
        
    request_phase_active = is_setting_true('request_phase_active')

    # Final advising courses sorting — resolve course codes from grade section_ids
    grades = Grade.query.filter_by(student_id=student.id).all()
    def _grade_course_code(g):
        sec = SectionOffering.query.get(g.section_id)
        return sec.course_code if sec else g.section_id

    passed_codes = [_grade_course_code(g) for g in grades if g.grade_letter not in ('F', None, '')]
    f_grades     = [_grade_course_code(g) for g in grades if g.grade_letter == 'F']
    d_grades     = [_grade_course_code(g) for g in grades if g.grade_letter in ['D', 'D+']]

    # Class Schedules
    current_sem_regs = Registration.query.filter_by(student_id=student.id, semester_id=get_current_semester()).all()
    current_schedule_sections = [SectionOffering.query.get(r.section_id) for r in current_sem_regs if SectionOffering.query.get(r.section_id)]
    current_course_codes = [sec.course_code for sec in current_schedule_sections if sec]
    all_completed_or_ongoing = set(passed_codes) | set(current_course_codes)

    # Recommended courses (prerequisites met, not yet passed or ongoing, and completed credit requirement is met)
    recommended_courses = []
    for c in courses:
        if c.code not in all_completed_or_ongoing:
            prereqs_met = all(pre in all_completed_or_ongoing for pre in c.prerequisites)
            credits_met = student.completed_credits >= c.completed_credit_requirement
            if prereqs_met and credits_met:
                recommended_courses.append(c.code)
                
    next_schedule_sections = [SectionOffering.query.get(r.section_id) for r in registrations if SectionOffering.query.get(r.section_id)]

    # Ledger
    ledger = LedgerEntry.query.filter_by(student_id=student.id).all()
    installments = Installment.query.filter_by(student_id=student.id).all()
    requests = AdvisingRequest.query.filter_by(student_id=student.id, semester_id=get_next_semester()).all()
    semester_drop_requests = SemesterDropRequest.query.filter_by(
        student_id=student.id,
        semester_id=get_next_semester()
    ).order_by(SemesterDropRequest.created_at.desc()).all()
    all_windows = AdvisingWindow.query.filter_by(semester_id=get_current_semester()).all()
    
    # Find student's matched advising windows based on completed credits
    student_pre_window = next((w for w in all_windows if w.type == 'pre' and w.credit_min <= student.completed_credits <= w.credit_max), None)
    student_final_window = next((w for w in all_windows if w.type == 'final' and w.credit_min <= student.completed_credits <= w.credit_max), None)

    now_iso = datetime.now().isoformat()
    
    def format_window_time(dt_str):
        if not dt_str:
            return 'Not Set Yet'
        try:
            t_str = dt_str.replace('T', ' ')
            if len(t_str) > 16:
                t_str = t_str[:16]
            dt = datetime.strptime(t_str, '%Y-%m-%d %H:%M')
            return dt.strftime('%b %d, %Y at %I:%M %p')
        except Exception:
            return dt_str.replace('T', ' ')

    def get_window_status(w, active_toggle):
        if not w:
            return 'Not Set Yet', 'text-slate-400 bg-slate-500/10'
        if w.start_date_time <= now_iso <= w.end_date_time:
            if active_toggle:
                return 'OPEN NOW', 'text-emerald-500 bg-emerald-500/10 animate-pulse font-extrabold border border-emerald-500/20'
            else:
                return 'Inactive (Toggled Off)', 'text-amber-500 bg-amber-500/10 font-bold border border-amber-500/20'
        elif now_iso < w.start_date_time:
            return 'Scheduled / Upcoming', 'text-blue-500 bg-blue-500/10 font-bold border border-blue-500/20'
        else:
            return 'Closed', 'text-rose-500 bg-rose-500/10 font-bold border border-rose-500/20'

    pre_status_label, pre_status_class = get_window_status(student_pre_window, pre_advising_active)
    final_status_label, final_status_class = get_window_status(student_final_window, final_advising_active)

    student_pre_start = format_window_time(student_pre_window.start_date_time) if student_pre_window else 'Not Set Yet'
    student_pre_end = format_window_time(student_pre_window.end_date_time) if student_pre_window else 'Not Set Yet'
    student_final_start = format_window_time(student_final_window.start_date_time) if student_final_window else 'Not Set Yet'
    student_final_end = format_window_time(student_final_window.end_date_time) if student_final_window else 'Not Set Yet'

    request_status_label = 'Closed'
    request_status_class = 'text-rose-500 bg-rose-500/10 font-bold border border-rose-500/20'
    if request_phase_active:
        if (len(plan_course_ids) > 0) and (not pre_advising_active) and (not final_advising_active):
            request_status_label = 'OPEN NOW'
            request_status_class = 'text-indigo-500 bg-indigo-500/10 animate-pulse font-extrabold border border-indigo-500/20'
        else:
            request_status_label = 'Locked (Pre-requisites pending)'
            request_status_class = 'text-amber-500 bg-amber-500/10 font-bold border border-amber-500/20'

    request_card_active = (len(plan_course_ids) > 0) and (not pre_advising_active) and (not final_advising_active) and request_phase_active
    
    return render_template(
        'student.html',
        active_tab=active_tab,
        request_card_active=request_card_active,
        student=student,
        dept=dept,
        advisor=advisor,
        total_required=total_required,
        remaining_credits=remaining_credits,
        announcements=announcements,
        courses=courses,
        courses_json=courses_json,
        plan_course_ids=plan_course_ids,
        plan_credits=plan_credits,
        sections=sections,
        sections_json=sections_json,
        selected_section_ids=selected_section_ids,
        registrations=registrations,
        pre_advising_active=pre_advising_active,
        final_advising_active=final_advising_active,
        request_phase_active=request_phase_active,
        f_grades=f_grades,
        d_grades=d_grades,
        recommended_courses=recommended_courses,
        current_schedule_sections=current_schedule_sections,
        next_schedule_sections=next_schedule_sections,
        grades=grades,
        ledger=ledger,
        installments=installments,
        requests=requests,
        semester_drop_requests=semester_drop_requests,
        all_windows=all_windows,
        advisor_email=advisor_email,
        pre_status_label=pre_status_label,
        pre_status_class=pre_status_class,
        final_status_label=final_status_label,
        final_status_class=final_status_class,
        student_pre_start=student_pre_start,
        student_pre_end=student_pre_end,
        student_final_start=student_final_start,
        student_final_end=student_final_end,
        request_status_label=request_status_label,
        request_status_class=request_status_class,
        datetime=datetime.now().strftime('%d-%b-%Y %I:%M %p')
    )

# Student saves pre-advising plan
@app.route('/student/save-plan', methods=['POST'])
@login_required
def save_plan():
    if current_user.role != 'student':
        if request.is_json:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
        return redirect(url_for('home'))
        
    student = Student.query.filter_by(user_id=current_user.id).first()
    if not student:
        if request.is_json:
            return jsonify({'status': 'error', 'message': 'Student profile not found.'}), 404
        flash("Student profile not found.", "error")
        return redirect(url_for('login_page'))
        
    # Gating checks - time window check
    now_str = datetime.now().strftime('%Y-%m-%dT%H:%M')
    semester = get_current_semester()
    window = AdvisingWindow.query.filter(
        AdvisingWindow.type == 'pre',
        AdvisingWindow.semester_id == semester,
        AdvisingWindow.credit_min <= student.completed_credits,
        AdvisingWindow.credit_max >= student.completed_credits,
        AdvisingWindow.start_date_time <= now_str,
        AdvisingWindow.end_date_time >= now_str
    ).first()
    if not window:
        msg = 'Pre-advising is currently closed or not active for your completed credits range.'
        if request.is_json:
            return jsonify({'status': 'error', 'message': msg})
        flash(msg, 'error')
        return redirect('/advising')
        
    if request.is_json:
        submitted_course_ids = request.json.get('course_ids', [])
    else:
        course_ids_str = request.form.get('course_ids', '[]')
        submitted_course_ids = json.loads(course_ids_str)
    
    # Enforce maximum courses: 6
    if len(submitted_course_ids) > 6:
        msg = 'Pre-advising constraint: Maximum of 6 courses allowed.'
        if request.is_json:
            return jsonify({'status': 'error', 'message': msg})
        flash(msg, 'error')
        return redirect('/advising')
        
    # Eligibility check on submitted courses
    eligible_course_codes = get_eligible_courses_for_student(student)
    for ccode in submitted_course_ids:
        if ccode not in eligible_course_codes:
            msg = f"Pre-advising constraint: You are not eligible to take course {ccode} due to prerequisite or completed credit requirements."
            if request.is_json:
                return jsonify({'status': 'error', 'message': msg})
            flash(msg, "error")
            return redirect('/advising')
            
    # Auto-expand to include eligible labs
    expanded_course_ids = []
    for ccode in submitted_course_ids:
        expanded_course_ids.append(ccode)
        lab_code = ccode + " Lab"
        if lab_code in eligible_course_codes:
            expanded_course_ids.append(lab_code)
    course_ids = list(set(expanded_course_ids))
    
    # Check credit limit on expanded list
    courses = PreAdvisingCourse.query.filter(PreAdvisingCourse.code.in_(course_ids)).all()
    total_credits = sum([c.credits for c in courses])
    
    # Enforce maximum credits: 15.0
    if total_credits > 15.0:
        msg = 'Pre-advising constraint: Total credits cannot exceed 15.0 CR.'
        if request.is_json:
            return jsonify({'status': 'error', 'message': msg})
        flash(msg, 'error')
        return redirect('/advising')
        
    plan = AdvisingPlan.query.filter_by(student_id=student.id, semester_id=get_next_semester()).first()
    if not plan:
        plan = AdvisingPlan(id='PLAN-'+student.id+'-'+get_next_semester(), student_id=student.id, semester_id=get_next_semester())
        db.session.add(plan)
        
    plan.course_ids = course_ids
    student.advising_status = 'planned'
    db.session.commit()
    
    if request.is_json:
        return jsonify({'status': 'success', 'message': 'Pre-advising plan auto-saved!'})
        
    flash('Pre-advising plan saved successfully!', 'success')
    return redirect('/advising')

# Dynamic section toggling (Auto-books/Auto-saves instantly)
@app.route('/student/toggle-section', methods=['POST'])
@login_required
def toggle_section():
    if current_user.role != 'student':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
        
    student = Student.query.filter_by(user_id=current_user.id).first()
    if not student:
        return jsonify({'status': 'error', 'message': 'Student profile not found.'}), 404
    
    # Gating checks - time window check
    now_str = datetime.now().strftime('%Y-%m-%dT%H:%M')
    semester = get_current_semester()
    window = AdvisingWindow.query.filter(
        AdvisingWindow.type == 'final',
        AdvisingWindow.semester_id == semester,
        AdvisingWindow.credit_min <= student.completed_credits,
        AdvisingWindow.credit_max >= student.completed_credits,
        AdvisingWindow.start_date_time <= now_str,
        AdvisingWindow.end_date_time >= now_str
    ).first()
    if not window:
        return jsonify({'status': 'error', 'message': 'Final advising is currently closed or not active for your completed credits range.'})
        
    # Financial hold check
    if not student.financial_cleared or student.outstanding_balance > 0:
        return jsonify({'status': 'error', 'message': 'Registration blocked due to outstanding financial balance.'})
        

        
    sec_id = request.json.get('section_id')
    sec = SectionOffering.query.get(sec_id)
    if not sec:
        return jsonify({'status': 'error', 'message': 'Section offering not found.'})
        
    existing_reg = Registration.query.filter_by(student_id=student.id, section_id=sec.id, semester_id=get_next_semester()).first()
    if not existing_reg:
        # Eligibility check
        eligible_course_codes = get_eligible_courses_for_student(student)
        if sec.course_code not in eligible_course_codes:
            return jsonify({'status': 'error', 'message': f"You are not eligible to register for course {sec.course_code} due to prerequisite or completed credit requirements."})
            
    if existing_reg:
        # Drop logic
        # If dropping a theory that has a linked lab, drop lab too
        db.session.delete(existing_reg)
        sec.enrolled_count = max(0, sec.enrolled_count - 1)
        
        if sec.linked_section_id:
            lab_reg = Registration.query.filter_by(student_id=student.id, section_id=sec.linked_section_id, semester_id=get_next_semester()).first()
            if lab_reg:
                db.session.delete(lab_reg)
                lab_sec = SectionOffering.query.get(sec.linked_section_id)
                if lab_sec:
                    lab_sec.enrolled_count = max(0, lab_sec.enrolled_count - 1)
                    
        db.session.commit()
        return jsonify({'status': 'success', 'action': 'dropped', 'message': f"Dropped {sec.course_code}."})
        
    # Add logic
    # Find active cart items
    regs = Registration.query.filter_by(student_id=student.id, semester_id=get_next_semester()).all()
    current_sections = [SectionOffering.query.get(r.section_id) for r in regs if SectionOffering.query.get(r.section_id)]
    
    # Enforce maximum constraints: 15.0 credits & 6 courses
    courses_count = len(current_sections) + 1
    total_credits = sum([s.credits for s in current_sections]) + sec.credits
    
    # If lab auto-added, account for its credit and count
    linked_lab = None
    if sec.linked_section_id:
        linked_lab = SectionOffering.query.get(sec.linked_section_id)
        if linked_lab:
            courses_count += 1
            total_credits += linked_lab.credits
            
    if courses_count > 6 or total_credits > 15.0:
        return jsonify({'status': 'error', 'message': 'Registration limit exceeded (Max 15.0 credits & 6 courses).'})
        
    # Capacity Check
    if sec.enrolled_count >= sec.capacity:
        return jsonify({'status': 'error', 'message': f"Section {sec.section_number} is full."})
    if linked_lab and linked_lab.enrolled_count >= linked_lab.capacity:
        return jsonify({'status': 'error', 'message': f"Linked section {linked_lab.section_number} is full."})
        
    # Department restriction check
    if sec.dedicated_departments and len(sec.dedicated_departments) > 0 and 'None' not in sec.dedicated_departments:
        if student.department_id not in sec.dedicated_departments:
            return jsonify({'status': 'error', 'message': f"Restricted to: {', '.join(sec.dedicated_departments)}."})
            
    # Schedule conflict check
    for cs in current_sections:
        if schedules_conflict(cs.schedule, sec.schedule):
            return jsonify({'status': 'error', 'message': f"Conflict with {cs.course_code} ({cs.schedule})."})
        if linked_lab and schedules_conflict(cs.schedule, linked_lab.schedule):
            return jsonify({'status': 'error', 'message': f"Linked Lab conflict with {cs.course_code} ({cs.schedule})."})
            
    # Register!
    reg = Registration(id=f"REG-{student.id}-{sec.id}", student_id=student.id, section_id=sec.id, semester_id=get_next_semester())
    db.session.add(reg)
    sec.enrolled_count += 1
    
    if linked_lab:
        existing_linked_reg = Registration.query.filter_by(student_id=student.id, section_id=linked_lab.id, semester_id=get_next_semester()).first()
        if not existing_linked_reg:
            lab_reg = Registration(id=f"REG-{student.id}-{linked_lab.id}", student_id=student.id, section_id=linked_lab.id, semester_id=get_next_semester())
            db.session.add(lab_reg)
            linked_lab.enrolled_count += 1
        
    student.advising_status = 'approved'
    db.session.commit()
    
    return jsonify({'status': 'success', 'action': 'registered', 'message': f"Enrolled in {sec.course_code}."})

# Submit advising override request (Add Course Option - Max 2 requests)
@app.route('/student/submit-request', methods=['POST'])
@login_required
def submit_override_request():
    if current_user.role != 'student':
        return redirect(url_for('home'))
        
    student = Student.query.filter_by(user_id=current_user.id).first()
    if not student:
        flash("Student profile not found.", "error")
        return redirect(url_for('login_page'))
    sec_id = (request.form.get('section_id') or '').strip()
    course_id = (request.form.get('course_id') or '').strip()
    comments = (request.form.get('comments') or '').strip()
    request_type = (request.form.get('request_type') or 'add_course').strip()

    if request_type not in ['add_course', 'seat_increase', 'drop_course']:
        request_type = 'add_course'

    active_setting = SystemSetting.query.filter_by(key='request_phase_active').first()
    if not active_setting or active_setting.value != 'true':
        flash('Course request phase is currently closed by the Administrator.', 'error')
        return redirect('/advising')

    if not comments:
        flash('Justification comments are required.', 'error')
        return redirect('/advising')

    if request_type == 'seat_increase':
        sec = SectionOffering.query.get(sec_id)
        if not sec:
            flash('Please select a valid section for the seat increase request.', 'error')
            return redirect('/advising')
        course_id = sec.course_code
    elif not course_id:
        flash('Please select a course before submitting the request.', 'error')
        return redirect('/advising')
    
    # Check max 2 requests
    req_count = AdvisingRequest.query.filter_by(student_id=student.id, semester_id=get_next_semester()).count()
    if req_count >= 2:
        flash('Maximum request limit reached (Max 2 requests allowed).', 'error')
        return redirect('/advising')
        
    if not student.advisor_id:
        flash('No advisor assigned to route request.', 'error')
        return redirect('/advising')
        
    req = AdvisingRequest(
        id=f"REQ-{student.id}-{int(datetime.utcnow().timestamp())}-{random.randint(1000, 9999)}",
        student_id=student.id,
        section_id=sec_id or None,
        course_id=course_id,
        type=request_type,
        comments=comments,
        semester_id=get_next_semester(),
        advisor_id=student.advisor_id
    )
    db.session.add(req)
    db.session.commit()
    
    flash('Request submitted to your advisor successfully.', 'success')
    return redirect('/advising')

# Submit Change Section Request
@app.route('/student/submit-change-request', methods=['POST'])
@login_required
def submit_change_request():
    if current_user.role != 'student':
        return redirect(url_for('home'))
        
    student = Student.query.filter_by(user_id=current_user.id).first()
    if not student:
        flash("Student profile not found.", "error")
        return redirect(url_for('login_page'))
    current_sec_id = (request.form.get('current_section_id') or '').strip()
    new_sec_id = (request.form.get('new_section_id') or '').strip()
    comments = (request.form.get('comments') or '').strip()

    active_setting = SystemSetting.query.filter_by(key='request_phase_active').first()
    if not active_setting or active_setting.value != 'true':
        flash('Course request phase is currently closed by the Administrator.', 'error')
        return redirect('/advising')

    if not student.advisor_id:
        flash('No advisor assigned to route request.', 'error')
        return redirect('/advising')

    if not comments:
        flash('Justification comments are required.', 'error')
        return redirect('/advising')
    
    req_count = AdvisingRequest.query.filter_by(student_id=student.id, semester_id=get_next_semester()).count()
    if req_count >= 2:
        flash('Maximum request limit reached (Max 2 requests allowed).', 'error')
        return redirect('/advising')
        
    current_sec = SectionOffering.query.get(current_sec_id)
    new_sec = SectionOffering.query.get(new_sec_id)
    
    if not current_sec or not new_sec:
        flash('Invalid sections selected.', 'error')
        return redirect('/advising')

    if current_sec.course_code.replace(' Lab', '') != new_sec.course_code.replace(' Lab', ''):
        flash('Section change must stay within the same course.', 'error')
        return redirect('/advising')
        
    # Check conflict with other courses in cart (excluding current section)
    regs = Registration.query.filter_by(student_id=student.id, semester_id=get_next_semester()).all()
    for r in regs:
        if r.section_id != current_sec_id:
            other = SectionOffering.query.get(r.section_id)
            if other and schedules_conflict(other.schedule, new_sec.schedule):
                flash(f"Conflict detected: Swapped section conflicts with {other.course_code}.", 'error')
                return redirect('/advising')
                
    req = AdvisingRequest(
        id=f"REQ-{student.id}-{int(datetime.utcnow().timestamp())}-{random.randint(1000, 9999)}",
        student_id=student.id,
        section_id=new_sec_id,
        current_section_id=current_sec_id,
        course_id=new_sec.course_code,
        type='change_section',
        comments=comments,
        semester_id=get_next_semester(),
        advisor_id=student.advisor_id
    )
    db.session.add(req)
    db.session.commit()
    
    flash('Section change swap request sent to academic advisor.', 'success')
    return redirect('/advising')

@app.route('/student/semester-drop', methods=['POST'])
@login_required
def submit_semester_drop():
    if current_user.role != 'student':
        return redirect(url_for('home'))

    student = Student.query.filter_by(user_id=current_user.id).first()
    if not student:
        flash('Student profile was not found.', 'error')
        return redirect(url_for('home'))

    reason = request.form.get('reason', '').strip()
    if len(reason) < 20:
        flash('Please provide a reason of at least 20 characters for semester drop.', 'error')
        return redirect('/advising')
    if len(reason) > 1000:
        flash('Semester drop reason must be 1000 characters or fewer.', 'error')
        return redirect('/advising')

    existing_pending = SemesterDropRequest.query.filter_by(
        student_id=student.id,
        semester_id=get_current_semester(),
        status='pending'
    ).first()
    if existing_pending:
        flash('You already have a pending semester drop request for this semester.', 'error')
        return redirect('/advising')

    drop_request = SemesterDropRequest(
        id=f"DROP-{student.id}-{int(datetime.utcnow().timestamp())}-{random.randint(1000, 9999)}",
        student_id=student.id,
        semester_id=get_current_semester(),
        reason=reason,
        status='pending'
    )
    db.session.add(drop_request)
    db.session.commit()

    flash('Semester drop application submitted successfully and marked as pending.', 'success')
    return redirect('/advising')

# FACULTY VIEWS
@app.route('/faculty')
@login_required
def faculty_dashboard():
    if current_user.role != 'faculty':
        return redirect(url_for('home'))
        
    faculty = current_faculty_profile()
    if not faculty:
        flash('Faculty profile was not found for this account.', 'error')
        return redirect(url_for('logout'))

    dept = Department.query.get(faculty.department_id)
    advisees = Student.query.filter_by(advisor_id=faculty.id).all()
    sections = SectionOffering.query.filter_by(
        faculty_id=faculty.id,
        semester_id=get_current_semester()
    ).order_by(SectionOffering.course_code.asc(), SectionOffering.section_number.asc()).all()
    
    # Course Requests FIFO
    requests = db.session.query(
        AdvisingRequest.id, AdvisingRequest.section_id, AdvisingRequest.current_section_id,
        AdvisingRequest.course_id, AdvisingRequest.type, AdvisingRequest.status,
        AdvisingRequest.comments, AdvisingRequest.advisor_note, AdvisingRequest.created_at,
        Student.name.label('student_name'), Student.id.label('student_id')
    ).join(Student, Student.id == AdvisingRequest.student_id)\
     .filter(AdvisingRequest.advisor_id == faculty.id, AdvisingRequest.semester_id == get_current_semester())\
     .order_by(AdvisingRequest.created_at.asc()).all()
     
    all_sections = SectionOffering.query.filter_by(
        semester_id=get_current_semester()
    ).order_by(SectionOffering.course_code.asc(), SectionOffering.section_number.asc()).all()
    dept_students = Student.query.filter_by(department_id=faculty.department_id).all()
     
    # Class Schedules
    prev_semester = get_previous_semester()
    
    # Pending request count for sidebar badge
    pending_count = sum(1 for r in requests if r.status == 'pending_advisor')

    # Format advisees JSON with computed remaining_credits
    advisees_json = json.dumps([{
        'id': s.id,
        'name': s.name,
        'cgpa': s.cgpa,
        'completed_credits': s.completed_credits,
        'remaining_credits': max(0.0, MAJOR_CREDITS.get(s.department_id, 140.0) - s.completed_credits),
        'has_plan': bool(AdvisingPlan.query.filter_by(student_id=s.id, semester_id=get_current_semester()).first())
    } for s in advisees])

    # Advisee pre-advising plan lookup set
    advisee_ids = [s.id for s in advisees]
    advisee_plans = AdvisingPlan.query.filter(
        AdvisingPlan.student_id.in_(advisee_ids),
        AdvisingPlan.semester_id == get_current_semester()
    ).all() if advisee_ids else []
    plan_student_ids = {p.student_id for p in advisee_plans}

    schedule_items = []
    for section in sections:
        schedule_items.extend(expand_section_schedule(section))
    schedule_items.sort(key=lambda item: (item['day_sort'], item['start_sort'], item['section'].course_code))

    sections_json = json.dumps([section_to_dict(s) for s in sections])
    all_sections_json = json.dumps([section_to_dict(s) for s in all_sections])

    return render_template(
        'faculty.html',
        faculty=faculty,
        dept=dept,
        advisees=advisees,
        advisees_json=advisees_json,
        sections=sections,
        all_sections=all_sections,
        requests=requests,
        pending_count=pending_count,
        dept_students=dept_students,
        prev_semester=prev_semester,
        dept_major_credits=MAJOR_CREDITS,
        plan_student_ids=plan_student_ids,
        schedule_items=schedule_items,
        sections_json=sections_json,
        all_sections_json=all_sections_json,
        active_semester=get_current_semester()
    )

# Faculty manually adds/drops courses for ANY student in department (All Student Advising)
@app.route('/faculty/advise-student', methods=['POST'])
@login_required
def faculty_advise_student():
    if current_user.role != 'faculty':
        return redirect(url_for('home'))
        
    std_id = request.form.get('student_id')
    action = request.form.get('action') # 'add' or 'drop'
    sec_id = request.form.get('section_id')
    
    student = Student.query.get(std_id)
    faculty = current_faculty_profile()
    
    if not faculty:
        flash('Faculty profile was not found.', 'error')
        return redirect(url_for('faculty_dashboard'))

    if not student or student.department_id != faculty.department_id:
        flash('Student record not found in your department.', 'error')
        return redirect(url_for('faculty_dashboard'))
        
    sec = SectionOffering.query.get(sec_id)
    if not sec:
        flash('Section offering not found.', 'error')
        return redirect(url_for('faculty_dashboard'))
        
    if action == 'add':
        existing_reg = Registration.query.filter_by(
            student_id=student.id,
            section_id=sec.id,
            semester_id=get_current_semester()
        ).first()
        if existing_reg:
            flash(f'{student.name} is already registered in this section.', 'error')
            return redirect(url_for('faculty_dashboard'))

        if sec.enrolled_count >= sec.capacity:
            flash('Error: Section is full.', 'error')
            return redirect(url_for('faculty_dashboard'))

        conflict_section = student_has_schedule_conflict(student.id, sec)
        if conflict_section:
            flash(f"Schedule conflict with {conflict_section.course_code} ({conflict_section.schedule}).", 'error')
            return redirect(url_for('faculty_dashboard'))
            
        reg = Registration(id=f"REG-{student.id}-{sec.id}", student_id=student.id, section_id=sec.id, semester_id=get_current_semester())
        db.session.add(reg)
        sec.enrolled_count += 1
        db.session.commit()
        flash(f"Successfully registered {sec.course_code} Sec {sec.section_number} for Student {student.name}.", 'success')
        
    elif action == 'drop':
        reg = Registration.query.filter_by(student_id=student.id, section_id=sec.id, semester_id=get_current_semester()).first()
        if reg:
            db.session.delete(reg)
            sec.enrolled_count = max(0, sec.enrolled_count - 1)
            db.session.commit()
            flash(f"Successfully dropped {sec.course_code} for Student {student.name}.", 'success')
            
    return redirect(url_for('faculty_dashboard'))

# Manual section change route (bypasses constraints on override)
@app.route('/faculty/swap-section', methods=['POST'])
@login_required
def faculty_swap_section():
    if current_user.role != 'faculty':
        return redirect(url_for('home'))
        
    std_id = request.form.get('student_id')
    current_sec_id = request.form.get('current_section_id')
    wanted_sec_id = request.form.get('wanted_section_id')
    
    student = Student.query.get(std_id)
    faculty = current_faculty_profile()
    if not faculty:
        flash('Faculty profile was not found.', 'error')
        return redirect(url_for('faculty_dashboard'))

    if not student or student.department_id != faculty.department_id:
        flash('Student not found.', 'error')
        return redirect(url_for('faculty_dashboard'))
        
    cur_reg = Registration.query.filter_by(student_id=student.id, section_id=current_sec_id, semester_id=get_current_semester()).first()
    current_sec = SectionOffering.query.get(current_sec_id)
    wanted_sec = SectionOffering.query.get(wanted_sec_id)
    
    if not cur_reg or not current_sec or not wanted_sec:
        flash('Invalid swap selections.', 'error')
        return redirect(url_for('faculty_dashboard'))

    if current_sec.course_code.replace(' Lab', '') != wanted_sec.course_code.replace(' Lab', ''):
        flash('Section swap must stay within the same course.', 'error')
        return redirect(url_for('faculty_dashboard'))

    existing_wanted = Registration.query.filter_by(
        student_id=student.id,
        section_id=wanted_sec.id,
        semester_id=get_current_semester()
    ).first()
    if existing_wanted:
        flash('Student is already registered in the target section.', 'error')
        return redirect(url_for('faculty_dashboard'))

    if wanted_sec.enrolled_count >= wanted_sec.capacity:
        flash('Target section is already full.', 'error')
        return redirect(url_for('faculty_dashboard'))

    conflict_section = student_has_schedule_conflict(student.id, wanted_sec, exclude_section_id=current_sec_id)
    if conflict_section:
        flash(f"Target section conflicts with {conflict_section.course_code} ({conflict_section.schedule}).", 'error')
        return redirect(url_for('faculty_dashboard'))
        
    # Process Swap
    db.session.delete(cur_reg)
    current_sec.enrolled_count = max(0, current_sec.enrolled_count - 1)
        
    new_reg = Registration(id=f"REG-{student.id}-{wanted_sec.id}", student_id=student.id, section_id=wanted_sec.id, semester_id=get_current_semester())
    db.session.add(new_reg)
    wanted_sec.enrolled_count += 1
    
    db.session.commit()
    flash(f"Swapped section successfully for {student.name}.", 'success')
    return redirect(url_for('faculty_dashboard'))

@app.route('/faculty/get-student-sections/<std_id>')
@login_required
def get_student_sections(std_id):
    if current_user.role != 'faculty':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
    faculty = current_faculty_profile()
    student = Student.query.get(std_id)
    if not faculty or not student or student.department_id != faculty.department_id:
        return jsonify({'status': 'error', 'message': 'Student is outside your department'}), 403

    regs = Registration.query.filter_by(student_id=std_id, semester_id=get_current_semester()).all()
    section_ids = [r.section_id for r in regs]
    details = []
    for reg in regs:
        section = SectionOffering.query.get(reg.section_id)
        if section:
            details.append(section_to_dict(section))
    return jsonify({'status': 'success', 'sections': section_ids, 'section_details': details})

@app.route('/faculty/section-roster/<sec_id>')
@login_required
def faculty_section_roster(sec_id):
    if current_user.role != 'faculty':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

    faculty = current_faculty_profile()
    section = SectionOffering.query.get(sec_id)
    if not faculty_can_teach_section(faculty, section):
        return jsonify({'status': 'error', 'message': 'Section is not assigned to you'}), 403

    attendance_date = request.args.get('date') or datetime.utcnow().date().isoformat()
    students = registered_students_for_section(section.id)
    attendance_rows = AttendanceRecord.query.filter_by(
        section_id=section.id,
        semester_id=get_current_semester(),
        date=attendance_date
    ).all()
    attendance_map = {row.student_id: row.status for row in attendance_rows}

    grade_rows = Grade.query.filter_by(
        section_id=section.course_code,
        semester_id=get_current_semester()
    ).all()
    grade_map = {row.student_id: row.grade_letter for row in grade_rows}

    roster = [{
        'id': student.id,
        'name': student.name,
        'cgpa': student.cgpa,
        'completed_credits': student.completed_credits,
        'attendance_status': attendance_map.get(student.id, 'present'),
        'grade_letter': grade_map.get(student.id, '')
    } for student in students]

    return jsonify({
        'status': 'success',
        'section': section_to_dict(section),
        'date': attendance_date,
        'students': roster
    })

@app.route('/faculty/save-attendance', methods=['POST'])
@login_required
def faculty_save_attendance():
    if current_user.role != 'faculty':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

    data = request.get_json(silent=True) or {}
    section = SectionOffering.query.get(data.get('section_id'))
    faculty = current_faculty_profile()
    if not faculty_can_teach_section(faculty, section):
        return jsonify({'status': 'error', 'message': 'Section is not assigned to you'}), 403

    attendance_date = data.get('date') or datetime.utcnow().date().isoformat()
    allowed_statuses = {'present', 'absent', 'late', 'excused'}
    registered_ids = {student.id for student in registered_students_for_section(section.id)}
    saved_count = 0

    for item in data.get('records', []):
        student_id = str(item.get('student_id', '')).strip()
        status = str(item.get('status', 'present')).strip().lower()
        if student_id not in registered_ids or status not in allowed_statuses:
            continue

        record_id = f"ATT-{attendance_date}-{section.id}-{student_id}"
        record = AttendanceRecord.query.get(record_id)
        if not record:
            record = AttendanceRecord(
                id=record_id,
                student_id=student_id,
                section_id=section.id,
                semester_id=get_current_semester(),
                date=attendance_date,
                status=status,
                marked_by=faculty.id
            )
            db.session.add(record)
        else:
            record.status = status
            record.marked_by = faculty.id
        saved_count += 1

    db.session.commit()
    return jsonify({'status': 'success', 'message': f'Attendance saved for {saved_count} students.'})

@app.route('/faculty/save-grades', methods=['POST'])
@login_required
def faculty_save_grades():
    if current_user.role != 'faculty':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

    data = request.get_json(silent=True) or {}
    section = SectionOffering.query.get(data.get('section_id'))
    faculty = current_faculty_profile()
    if not faculty_can_teach_section(faculty, section):
        return jsonify({'status': 'error', 'message': 'Section is not assigned to you'}), 403

    registered_ids = {student.id for student in registered_students_for_section(section.id)}
    changed_students = set()

    for item in data.get('grades', []):
        student_id = str(item.get('student_id', '')).strip()
        grade_letter = str(item.get('grade_letter', '')).strip().upper()
        if student_id not in registered_ids:
            continue
        if save_grade_for_student(student_id, section.course_code, grade_letter, get_current_semester()):
            changed_students.add(student_id)

    db.session.commit()
    for student_id in changed_students:
        recalculate_student_stats(student_id)

    return jsonify({'status': 'success', 'message': f'Grades saved for {len(changed_students)} students.'})

@app.route('/faculty/submit-grade', methods=['POST'])
@login_required
def faculty_submit_grade():
    if current_user.role != 'faculty':
        return redirect(url_for('home'))
    student_id = request.form.get('student_id', '').strip()
    course_code = request.form.get('course_code', '').strip().upper()
    grade_letter = request.form.get('grade_letter', '').strip().upper()
    semester_id = request.form.get('semester_id', get_current_semester()).strip()
    if grade_letter not in GRADE_POINTS_MAP:
        flash(f"Invalid grade '{grade_letter}'.", 'error')
        return redirect(url_for('faculty_dashboard'))
    student = Student.query.get(student_id)
    if not student:
        flash('Student not found.', 'error')
        return redirect(url_for('faculty_dashboard'))
    if not save_grade_for_student(student_id, course_code, grade_letter, semester_id):
        flash(f"Invalid grade '{grade_letter}'.", 'error')
        return redirect(url_for('faculty_dashboard'))
    db.session.commit()
    recalculate_student_stats(student_id)
    flash(f"Grade '{grade_letter}' saved for {student_id} in {course_code}.", 'success')
    return redirect(url_for('faculty_dashboard'))

@app.route('/faculty/register-advisee-override', methods=['POST'])
@login_required
def register_advisee_override():
    if current_user.role != 'faculty':
        return redirect(url_for('home'))
        
    std_id = request.form.get('student_id')
    sec_ids_str = request.form.get('section_ids', '[]')
    try:
        sec_ids = json.loads(sec_ids_str)
    except json.JSONDecodeError:
        sec_ids = []
    
    student = Student.query.get(std_id)
    faculty = current_faculty_profile()
    if not faculty:
        flash('Faculty profile was not found.', 'error')
        return redirect(url_for('faculty_dashboard'))

    if not student or student.department_id != faculty.department_id:
        flash('Student record not found.', 'error')
        return redirect(url_for('faculty_dashboard'))
        
    # Drop all previous registrations
    old_regs = Registration.query.filter_by(student_id=student.id, semester_id=get_next_semester()).all()
    for reg in old_regs:
        sec = SectionOffering.query.get(reg.section_id)
        if sec:
            sec.enrolled_count = max(0, sec.enrolled_count - 1)
        db.session.delete(reg)
        
    # Register new selections
    registered_ids = set()
    for sid in sec_ids:
        if sid in registered_ids:
            continue
        sec = SectionOffering.query.get(sid)
        if not sec:
            continue
            
        reg = Registration(id=f"REG-{student.id}-{sec.id}", student_id=student.id, section_id=sec.id, semester_id=get_next_semester())
        db.session.add(reg)
        sec.enrolled_count += 1
        registered_ids.add(sec.id)
        
        # Auto add linked lab/theory
        if sec.linked_section_id and sec.linked_section_id not in registered_ids:
            linked_sec = SectionOffering.query.get(sec.linked_section_id)
            if linked_sec:
                l_reg = Registration(id=f"REG-{student.id}-{linked_sec.id}", student_id=student.id, section_id=linked_sec.id, semester_id=get_next_semester())
                db.session.add(l_reg)
                linked_sec.enrolled_count += 1
                registered_ids.add(linked_sec.id)
                
    db.session.commit()
    flash(f"Manual advising overrides saved successfully for {student.name}.", 'success')
    return redirect(url_for('faculty_dashboard'))

# Advisor action on student course requests
@app.route('/faculty/resolve-request/<req_id>', methods=['POST'])
@login_required
def resolve_request(req_id):
    if current_user.role != 'faculty':
        return redirect(url_for('home'))
        
    status = request.form.get('status') # 'approved' or 'rejected'
    advisor_note = request.form.get('advisor_note', '')
    req = AdvisingRequest.query.get(req_id)
    
    faculty = current_faculty_profile()
    if status not in ['approved', 'rejected']:
        flash("Invalid request decision.", "error")
        return redirect(url_for('faculty_dashboard'))

    if req and faculty and req.advisor_id == faculty.id:
        req.status = status
        req.advisor_note = advisor_note
        
        if status == 'approved':
            # Execute request actions in database
            student = Student.query.get(req.student_id)
            sec = find_requested_section(req)

            if not student or not sec:
                flash("Could not resolve the requested section. Ask the student to submit a valid section.", "error")
                return redirect(url_for('faculty_dashboard'))

            if req.type != 'seat_increase' and sec.enrolled_count >= sec.capacity:
                flash(f"Cannot approve: {sec.course_code} section {sec.section_number} is full.", "error")
                return redirect(url_for('faculty_dashboard'))
            
            if req.type == 'add_course' and sec:
                existing_reg = Registration.query.filter_by(
                    student_id=student.id,
                    section_id=sec.id,
                    semester_id=req.semester_id
                ).first()
                conflict_section = student_has_schedule_conflict(student.id, sec)
                if existing_reg:
                    flash("Student is already registered in that section.", "error")
                    return redirect(url_for('faculty_dashboard'))
                if conflict_section:
                    flash(f"Cannot approve: schedule conflict with {conflict_section.course_code}.", "error")
                    return redirect(url_for('faculty_dashboard'))

                linked_lab = None
                if sec.linked_section_id:
                    linked_lab = SectionOffering.query.get(sec.linked_section_id)

                if linked_lab:
                    lab_conflict = student_has_schedule_conflict(student.id, linked_lab)
                    if lab_conflict:
                        flash(f"Cannot approve: linked lab conflicts with {lab_conflict.course_code}.", "error")
                        return redirect(url_for('faculty_dashboard'))

                db.session.add(Registration(
                    id=f"REG-{student.id}-{sec.id}",
                    student_id=student.id,
                    section_id=sec.id,
                    semester_id=req.semester_id
                ))
                sec.enrolled_count += 1

                if linked_lab:
                    lab_reg = Registration.query.filter_by(student_id=student.id, section_id=linked_lab.id, semester_id=req.semester_id).first()
                    if not lab_reg:
                        db.session.add(Registration(
                            id=f"REG-{student.id}-{linked_lab.id}",
                            student_id=student.id,
                            section_id=linked_lab.id,
                            semester_id=req.semester_id
                        ))
                        linked_lab.enrolled_count += 1
                
            elif req.type == 'change_section' and sec:
                old_reg = Registration.query.filter_by(student_id=student.id, section_id=req.current_section_id, semester_id=req.semester_id).first()
                if not old_reg:
                    flash("Cannot approve: current registration was not found.", "error")
                    return redirect(url_for('faculty_dashboard'))

                target_reg = Registration.query.filter_by(
                    student_id=student.id,
                    section_id=sec.id,
                    semester_id=req.semester_id
                ).first()
                if target_reg and sec.id != req.current_section_id:
                    flash("Cannot approve: student is already registered in the target section.", "error")
                    return redirect(url_for('faculty_dashboard'))

                old_sec = SectionOffering.query.get(req.current_section_id)
                old_lab_id = old_sec.linked_section_id if old_sec else None

                conflict_section = student_has_schedule_conflict(student.id, sec, exclude_section_id=[req.current_section_id, old_lab_id])
                if conflict_section:
                    flash(f"Cannot approve: target section conflicts with {conflict_section.course_code}.", "error")
                    return redirect(url_for('faculty_dashboard'))

                linked_lab = None
                if sec.linked_section_id:
                    linked_lab = SectionOffering.query.get(sec.linked_section_id)

                if linked_lab:
                    lab_conflict = student_has_schedule_conflict(student.id, linked_lab, exclude_section_id=[req.current_section_id, old_lab_id])
                    if lab_conflict:
                        flash(f"Cannot approve: target lab section conflicts with {lab_conflict.course_code}.", "error")
                        return redirect(url_for('faculty_dashboard'))

                if old_reg:
                    db.session.delete(old_reg)
                    if old_sec:
                        old_sec.enrolled_count = max(0, old_sec.enrolled_count - 1)

                if old_lab_id:
                    old_lab_reg = Registration.query.filter_by(student_id=student.id, section_id=old_lab_id, semester_id=req.semester_id).first()
                    if old_lab_reg:
                        db.session.delete(old_lab_reg)
                        old_lab_sec = SectionOffering.query.get(old_lab_id)
                        if old_lab_sec:
                            old_lab_sec.enrolled_count = max(0, old_lab_sec.enrolled_count - 1)

                db.session.add(Registration(
                    id=f"REG-{student.id}-{sec.id}",
                    student_id=student.id,
                    section_id=sec.id,
                    semester_id=req.semester_id
                ))
                sec.enrolled_count += 1

                if linked_lab:
                    db.session.add(Registration(
                        id=f"REG-{student.id}-{linked_lab.id}",
                        student_id=student.id,
                        section_id=linked_lab.id,
                        semester_id=req.semester_id
                    ))
                    linked_lab.enrolled_count += 1

            elif req.type == 'seat_increase' and sec:
                sec.capacity += 1
                if sec.linked_section_id:
                    linked_sec = SectionOffering.query.get(sec.linked_section_id)
                    if linked_sec:
                        linked_sec.capacity += 1
                
        db.session.commit()
        flash(f"Request resolved as '{status}'.", 'success')
    else:
        flash("Request record not found or not assigned to you.", "error")
        
    return redirect(url_for('faculty_dashboard'))

# ADMIN VIEWS
@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    students = Student.query.all()
    faculties = Faculty.query.all()
    pre_courses = PreAdvisingCourse.query.all()
    section_offerings = SectionOffering.query.filter_by(semester_id=get_next_semester()).all()
    windows = AdvisingWindow.query.filter_by(semester_id=get_current_semester()).all()
    departments = Department.query.all()
    
    # Calculate Pre-Advising Demand
    demand_counts = {}
    plans = AdvisingPlan.query.filter_by(semester_id=get_next_semester()).all()
    for p in plans:
        for cCode in p.course_ids:
            demand_counts[cCode] = demand_counts.get(cCode, 0) + 1
            
    for pc in pre_courses:
        if pc.code not in demand_counts:
            demand_counts[pc.code] = 0
            
    demand_data = sorted(demand_counts.items(), key=lambda x: x[1], reverse=True)

    # Fetch settings
    def is_setting_true(key):
        s = SystemSetting.query.filter_by(key=key).first()
        return s.value == 'true' if s else False

    all_requests = AdvisingRequest.query.all()
    all_grades = Grade.query.all()
    
    # Build a map of user_id -> User for roster status display
    all_users = User.query.all()
    users_map = {u.id: u for u in all_users}
    
    # Build faculty map for advisor assignment display
    faculty_map = {f.id: f for f in faculties}

    student_map = {s.id: s for s in students}

    # Query 0-credit students
    students_0cr = Student.query.filter((Student.completed_credits == 0) | (Student.completed_credits == 0.0) | (Student.completed_credits == None)).all()
    for s in students_0cr:
        regs = Registration.query.filter_by(student_id=s.id, semester_id=get_next_semester()).all()
        s.registered_credits = sum([SectionOffering.query.get(r.section_id).credits for r in regs if SectionOffering.query.get(r.section_id)])

    return render_template(
        'admin.html',
        students=students,
        faculties=faculties,
        pre_courses=pre_courses,
        section_offerings=section_offerings,
        windows=windows,
        demand_data=demand_data,
        departments=departments,
        all_requests=all_requests,
        all_grades=all_grades,
        users_map=users_map,
        faculty_map=faculty_map,
        student_map=student_map,
        students_0cr=students_0cr,
        current_semester=get_current_semester(),
        pre_advising_active=is_setting_true('pre_advising_active'),
        final_advising_active=is_setting_true('final_advising_active'),
        request_phase_active=is_setting_true('request_phase_active')
    )

# Admin toggle settings
@app.route('/admin/toggle-setting/<key>', methods=['POST'])
@login_required
def toggle_setting(key):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    setting = SystemSetting.query.filter_by(key=key).first()
    if not setting:
        setting = SystemSetting(key=key, value='false')
        db.session.add(setting)
        
    setting.value = 'true' if setting.value == 'false' else 'false'
    db.session.commit()
    flash(f"System setting '{key}' toggled to {setting.value}.", 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/update-current-semester', methods=['POST'])
@login_required
def update_current_semester():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    curr_sem = request.form.get('current_semester')
    c_start = request.form.get('current_semester_start')
    c_end = request.form.get('current_semester_end')
    
    valid_sems = ['Spring2026', 'Summer2026', 'Fall2026', 'Spring2027']
    if curr_sem not in valid_sems:
        flash('Invalid Semester name.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=settings')
        
    def save_setting(key, val):
        setting = SystemSetting.query.filter_by(key=key).first()
        if not setting:
            setting = SystemSetting(key=key, value=val)
            db.session.add(setting)
        else:
            setting.value = val

    save_setting('current_semester', curr_sem)
    save_setting('current_semester_start', c_start)
    save_setting('current_semester_end', c_end)
    
    db.session.commit()
    flash('Current semester settings updated successfully!', 'success')
    return redirect(url_for('admin_dashboard') + '?tab=settings')

@app.route('/admin/update-next-semester', methods=['POST'])
@login_required
def update_next_semester():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    next_sem = request.form.get('next_semester')
    n_start = request.form.get('next_semester_start')
    n_end = request.form.get('next_semester_end')
    
    valid_sems = ['Spring2026', 'Summer2026', 'Fall2026', 'Spring2027']
    if next_sem not in valid_sems:
        flash('Invalid Semester name.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=settings')
        
    def save_setting(key, val):
        setting = SystemSetting.query.filter_by(key=key).first()
        if not setting:
            setting = SystemSetting(key=key, value=val)
            db.session.add(setting)
        else:
            setting.value = val

    save_setting('next_semester', next_sem)
    save_setting('next_semester_start', n_start)
    save_setting('next_semester_end', n_end)
    
    db.session.commit()
    flash('Next semester settings updated successfully!', 'success')
    return redirect(url_for('admin_dashboard') + '?tab=settings')

@app.route('/admin/perform-rollover', methods=['POST'])
@login_required
def perform_rollover():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    curr_sem = get_current_semester()
    next_sem = get_next_semester()
    
    # Rollover: Move registrations from next semester to current semester
    next_regs = Registration.query.filter_by(semester_id=next_sem).all()
    for reg in next_regs:
        reg.semester_id = next_sem # wait, old next_sem is the new current_sem!
        
    # Get sequence sequence
    sem_sequence = ['Spring2026', 'Summer2026', 'Fall2026', 'Spring2027']
    try:
        idx = sem_sequence.index(next_sem)
        new_next_sem = sem_sequence[idx + 1] if idx + 1 < len(sem_sequence) else 'Spring2027'
    except ValueError:
        new_next_sem = 'Spring2027'
        
    def save_setting(key, val):
        setting = SystemSetting.query.filter_by(key=key).first()
        if not setting:
            setting = SystemSetting(key=key, value=val)
            db.session.add(setting)
        else:
            setting.value = val
            
    next_start = SystemSetting.query.filter_by(key='next_semester_start').first()
    next_end = SystemSetting.query.filter_by(key='next_semester_end').first()
    
    save_setting('current_semester', next_sem)
    if next_start:
        save_setting('current_semester_start', next_start.value)
    if next_end:
        save_setting('current_semester_end', next_end.value)
        
    save_setting('next_semester', new_next_sem)
    save_setting('next_semester_start', '')
    save_setting('next_semester_end', '')
    
    db.session.commit()
    flash(f"Semester Rollover executed successfully! {next_sem} is now the Current Semester.", 'success')
    return redirect(url_for('admin_dashboard') + '?tab=settings')

@app.route('/admin/change-password', methods=['POST'])
@login_required
def change_admin_password():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    new_pass = request.form.get('password')
    if not new_pass or len(new_pass) < 4:
        flash('Password must be at least 4 characters long.', 'error')
        return redirect(url_for('admin_dashboard'))
        
    user = User.query.get(current_user.id)
    user.password_hash = generate_password_hash(new_pass)
    db.session.commit()
    
    flash('Admin password updated successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

def get_student_semester_level(student_id, current_semester_str):
    parts = student_id.split('-')
    if len(parts) < 3:
        return 1
    try:
        admit_year = int(parts[0])
        admit_term = int(parts[1])
    except ValueError:
        return 1
        
    import re
    match = re.match(r'([A-Za-z]+)(\d+)', current_semester_str)
    if not match:
        return 1
    term_name = match.group(1).lower()
    curr_year = int(match.group(2))
    
    term_map = {'spring': 1, 'summer': 2, 'fall': 3}
    curr_term = term_map.get(term_name, 1)
    
    elapsed = (curr_year - admit_year) * 3 + (curr_term - admit_term) + 1
    return max(1, elapsed)

def execute_default_advising(student):
    curr_sem = get_current_semester()
    # 1. Clear any existing registrations for this student in the current semester
    old_regs = Registration.query.filter_by(student_id=student.id, semester_id=curr_sem).all()
    for reg in old_regs:
        sec = SectionOffering.query.get(reg.section_id)
        if sec:
            sec.enrolled_count = max(0, sec.enrolled_count - 1)
        db.session.delete(reg)
    db.session.flush()
    
    # 2. Courses to assign based on student ID parsing
    parts = student.id.split('-')
    dept_code = parts[2] if len(parts) >= 3 else ''
    sem_level = get_student_semester_level(student.id, curr_sem)
    
    # Department codes mapping: 60 = CSE, 80 = EEE, 50 = ICE, 40 = ENG
    if dept_code == '60' and sem_level == 1:
        target_courses = ['CSE106', 'ENG101', 'MAT101']
    elif dept_code == '80' and sem_level == 1:
        target_courses = ['PHY109', 'CHE109', 'MAT101']
    elif dept_code == '50' and sem_level == 2:
        target_courses = ['ICE101', 'ENG101', 'MAT101']
    else:
        # Fallback default courses
        if dept_code == '60':
            target_courses = ['CSE103', 'ENG101', 'MAT101']
        elif dept_code == '80':
            target_courses = ['EEE101', 'MAT101', 'PHY109']
        elif dept_code == '50':
            target_courses = ['ICE101', 'ENG101', 'MAT101']
        else:
            target_courses = ['ENG101', 'MAT101', 'GEN226']
            
    for course_code in target_courses:
        # Check if already registered (e.g. from linked labs)
        already_reg = False
        for r in Registration.query.filter_by(student_id=student.id, semester_id=curr_sem).all():
            rsec = SectionOffering.query.get(r.section_id)
            if rsec and rsec.course_code == course_code:
                already_reg = True
                break
        if already_reg:
            continue
            
        sections = SectionOffering.query.filter_by(course_code=course_code, semester_id=curr_sem).all()
        if not sections:
            continue
            
        import random
        random.shuffle(sections)
        
        assigned = False
        for sec in sections:
            if sec.enrolled_count >= sec.capacity:
                continue
            if student_has_schedule_conflict(student.id, sec):
                continue
                
            # If it has a linked lab, check capacity & conflicts too
            linked_lab = None
            if sec.linked_section_id:
                linked_lab = SectionOffering.query.get(sec.linked_section_id)
                if linked_lab:
                    if linked_lab.enrolled_count >= linked_lab.capacity:
                        continue
                    if student_has_schedule_conflict(student.id, linked_lab):
                        continue
            
            # Found a free, non-conflicting section!
            reg = Registration(
                id=f"REG-{student.id}-{sec.id}",
                student_id=student.id,
                section_id=sec.id,
                semester_id=curr_sem
            )
            db.session.add(reg)
            sec.enrolled_count += 1
            
            # Register linked lab
            if linked_lab:
                reg_lab = Registration(
                    id=f"REG-{student.id}-{linked_lab.id}",
                    student_id=student.id,
                    section_id=linked_lab.id,
                    semester_id=curr_sem
                )
                db.session.add(reg_lab)
                linked_lab.enrolled_count += 1
                
            db.session.flush()
            assigned = True
            break
            
        if not assigned:
            raise ValueError(f"Could not find a conflict-free section with available seats for {course_code}.")
            
    student.advising_status = 'approved'
    db.session.commit()

@app.route('/admin/run-default-advising/<student_id>', methods=['POST'])
@login_required
def run_default_advising(student_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    student = Student.query.get(student_id)
    if not student:
        flash('Student not found.', 'error')
        return redirect(url_for('admin_dashboard'))
        
    try:
        execute_default_advising(student)
        flash(f"Default advising completed successfully for {student.name} ({student.id}).", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error advising {student.name}: {str(e)}", 'error')
        
    return redirect(url_for('admin_dashboard') + '?tab=default-advising')

@app.route('/admin/run-all-default-advising', methods=['POST'])
@login_required
def run_all_default_advising():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    # Get all students with 0 credits
    students_0cr = Student.query.filter((Student.completed_credits == 0) | (Student.completed_credits == 0.0) | (Student.completed_credits == None)).all()
    
    success_count = 0
    fail_count = 0
    errors = []
    
    for s in students_0cr:
        try:
            execute_default_advising(s)
            success_count += 1
        except Exception as e:
            db.session.rollback()
            fail_count += 1
            errors.append(f"{s.id}: {str(e)}")
            
    if fail_count == 0:
        flash(f"Default advising successfully completed for all {success_count} freshman students.", 'success')
    else:
        flash(f"Default advising completed with errors. Success: {success_count}, Failed: {fail_count}. Errors: {'; '.join(errors[:3])}", 'warning')
        
    return redirect(url_for('admin_dashboard') + '?tab=default-advising')

@app.route('/change-password', methods=['POST'])
@login_required
def change_password():
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()
    
    if not new_password or len(new_password) < 4:
        flash('Password must be at least 4 characters long.', 'error')
        return redirect(request.referrer or url_for('home'))
        
    if new_password != confirm_password:
        flash('Passwords do not match.', 'error')
        return redirect(request.referrer or url_for('home'))
        
    user = User.query.get(current_user.id)
    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    
    flash('Password updated successfully!', 'success')
    return redirect(request.referrer or url_for('home'))

@app.route('/admin/create-student', methods=['POST'])
@login_required
def create_student():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    std_id = request.form.get('student_id', '').strip()
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    dept = request.form.get('department_id')
    credits = float(request.form.get('completed_credits', 0.0))
    cgpa = float(request.form.get('cgpa', 0.0))
    balance = int(request.form.get('outstanding_balance', 0))
    cleared = 'financial_cleared' in request.form
    about = request.form.get('about', '').strip()
    
    # New fields
    phone_number = request.form.get('phone_number', '').strip()
    remaining_credits = float(request.form.get('remaining_credits', 0.0))
    present_address = request.form.get('present_address', '').strip()
    permanent_address = request.form.get('permanent_address', '').strip()
    completed_courses_and_grades = request.form.get('completed_courses_and_grades', '').strip()
    current_courses = request.form.get('current_courses', '').strip()
    current_course_credit = float(request.form.get('current_course_credit', 0.0))
    next_semester_courses = request.form.get('next_semester_courses', '').strip()
    next_semester_course_credit = float(request.form.get('next_semester_course_credit', 0.0))
    
    if len(about) > 500:
        flash('About section must be 500 characters or fewer.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=students')
    
    if not std_id or not name or not email:
        flash('All fields are required.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=students')
        
    if User.query.filter_by(email=email).first():
        flash('Email already registered.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=students')
        
    if Student.query.get(std_id):
        flash('Student ID already exists.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=students')
        
    try:
        profile_pic_filename = save_profile_pic_upload(request.files.get('profile_pic'), f"student_{std_id}")
    except ValueError as exc:
        flash(str(exc), 'error')
        return redirect(url_for('admin_dashboard') + '?tab=students')

    user = User(
        id='usr-' + std_id,
        email=email,
        password_hash=generate_password_hash('password123'),
        role='student',
        is_active=True,
        is_activated=False
    )
    db.session.add(user)
    db.session.flush()
    
    student = Student(
        id=std_id,
        user_id=user.id,
        name=name,
        department_id=dept,
        completed_credits=credits,
        cgpa=cgpa,
        outstanding_balance=balance,
        financial_cleared=cleared,
        profile_pic=profile_pic_filename,
        about=about,
        
        # Save new fields
        phone_number=phone_number if phone_number else None,
        remaining_credits=remaining_credits,
        present_address=present_address if present_address else None,
        permanent_address=permanent_address if permanent_address else None,
        completed_courses_and_grades=completed_courses_and_grades if completed_courses_and_grades else None,
        current_courses=current_courses if current_courses else None,
        current_course_credit=current_course_credit,
        next_semester_courses=next_semester_courses if next_semester_courses else None,
        next_semester_course_credit=next_semester_course_credit
    )
    db.session.add(student)
    
    # Sync grades and registrations
    import re
    if completed_courses_and_grades and completed_courses_and_grades.lower() != 'none':
        parts = re.split(r'[,;\n\r]+', completed_courses_and_grades)
        for part in parts:
            part = part.strip()
            if not part: continue
            match = re.search(r'([A-Za-z0-9\s]+)[:\s]+([A-Za-z+-]+)', part)
            if match:
                ccode = match.group(1).strip().replace(' ', '')
                gletter = match.group(2).strip()
                points_map = {
                    'A+': 4.0, 'A': 4.0, 'A-': 3.7,
                    'B+': 3.3, 'B': 3.0, 'B-': 2.7,
                    'C+': 2.3, 'C': 2.0, 'C-': 1.7,
                    'D+': 1.3, 'D': 1.0, 'F': 0.0
                }
                gpoint = points_map.get(gletter.upper(), 0.0)
                grade_id = f"GRD-{std_id}-{ccode}"
                new_grade = Grade(
                    id=grade_id,
                    student_id=std_id,
                    section_id=ccode,
                    grade_letter=gletter.upper(),
                    grade_point=gpoint,
                    semester_id='completed'
                )
                db.session.add(new_grade)
                
    if current_courses and current_courses.lower() != 'none':
        ccodes = [c.strip() for c in re.split(r'[,;\s]+', current_courses) if c.strip()]
        for cc in ccodes:
            if not cc: continue
            sec = SectionOffering.query.filter_by(course_code=cc).first()
            if sec:
                reg_id = f"REG-{std_id}-{sec.id}"
                existing_reg = Registration.query.get(reg_id)
                if not existing_reg:
                    new_reg = Registration(
                        id=reg_id,
                        student_id=std_id,
                        section_id=sec.id,
                        semester_id=sec.semester_id,
                        status='registered'
                    )
                    db.session.add(new_reg)
                    sec.enrolled_count = min(sec.capacity, sec.enrolled_count + 1)
                    
    db.session.commit()
    
    flash('Student account created successfully! Advise student to activate using their email.', 'success')
    return redirect(url_for('admin_dashboard') + '?tab=students')

@app.route('/admin/create-faculty', methods=['POST'])
@login_required
def create_faculty():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    fac_id = request.form.get('faculty_id', '').strip()
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    dept = request.form.get('department_id')
    about = request.form.get('about', '').strip()
    if len(about) > 500:
        flash('About section must be 500 characters or fewer.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    if not fac_id or not name or not email:
        flash('All fields are required.', 'error')
        return redirect(url_for('admin_dashboard'))
        
    if User.query.filter_by(email=email).first():
        flash('Email already registered.', 'error')
        return redirect(url_for('admin_dashboard'))
        
    if Faculty.query.get(fac_id):
        flash('Faculty ID already exists.', 'error')
        return redirect(url_for('admin_dashboard'))
        
    try:
        profile_pic_filename = save_profile_pic_upload(request.files.get('profile_pic'), f"faculty_{fac_id}")
    except ValueError as exc:
        flash(str(exc), 'error')
        return redirect(url_for('admin_dashboard'))

    user = User(
        id='usr-' + fac_id,
        email=email,
        password_hash=generate_password_hash('password123'),
        role='faculty',
        is_active=True,
        is_activated=False # Starts unactivated! First time activation flow.
    )
    db.session.add(user)
    
    faculty = Faculty(
        id=fac_id,
        user_id=user.id,
        name=name,
        department_id=dept,
        profile_pic=profile_pic_filename,
        about=about
    )
    db.session.add(faculty)
    db.session.commit()
    
    flash('Faculty account created successfully! Advise faculty to activate using their email.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/edit-capacity/<sec_id>', methods=['POST'])
@login_required
def edit_section_capacity(sec_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    cap = int(request.form.get('capacity', 30))
    sec = SectionOffering.query.get(sec_id)
    if sec:
        sec.capacity = cap
        # Synchronize linked section capacity if present
        if sec.linked_section_id:
            linked_sec = SectionOffering.query.get(sec.linked_section_id)
            if linked_sec:
                linked_sec.capacity = cap
        db.session.commit()
        flash(f"Section '{sec_id}' capacity updated to {cap}.", 'success')
    else:
        flash('Section offering not found.', 'error')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/toggle-status/<user_id>', methods=['POST'])
@login_required
def toggle_user_status(user_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    user = User.query.get(user_id)
    if user:
        user.is_active = not user.is_active
        db.session.commit()
        status_text = 'Activated' if user.is_active else 'Deactivated'
        flash(f"Portal for '{user.email}' is now {status_text}.", 'success')
    else:
        flash('User account not found.', 'error')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/auto-assign-advisors', methods=['POST'])
@login_required
def auto_assign_advisors():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    departments = Department.query.all()
    for dept in departments:
        dept_students = Student.query.filter_by(department_id=dept.id).all()
        dept_faculty = Faculty.query.filter_by(department_id=dept.id).all()
        
        if not dept_faculty or not dept_students:
            continue
            
        num_faculty = len(dept_faculty)
        num_students = len(dept_students)
        
        base_alloc = num_students // num_faculty
        remainder = num_students % num_faculty
        
        random.shuffle(dept_students)
        
        idx = 0
        for f_idx, faculty in enumerate(dept_faculty):
            count = base_alloc + (1 if f_idx < remainder else 0)
            for _ in range(count):
                dept_students[idx].advisor_id = faculty.id
                idx += 1
                 
    db.session.commit()
    flash('Advisors load-balanced and assigned randomly per department!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/assign-advisor', methods=['POST'])
@login_required
def assign_advisor():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    
    student_id = request.form.get('student_id', '').strip()
    faculty_id = request.form.get('faculty_id', '').strip()
    
    student = Student.query.get(student_id)
    if not student:
        flash('Student not found.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    if faculty_id:
        faculty = Faculty.query.get(faculty_id)
        if not faculty:
            flash('Faculty not found.', 'error')
            return redirect(url_for('admin_dashboard'))
        student.advisor_id = faculty_id
        flash(f'Advisor "{faculty.name}" assigned to {student.name} successfully.', 'success')
    else:
        student.advisor_id = None
        flash(f'Advisor unassigned from {student.name}.', 'success')
    
    db.session.commit()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete-student/<std_id>', methods=['POST'])
@login_required
def delete_student(std_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    student = Student.query.get(std_id)
    if student:
        user = User.query.get(student.user_id)
        db.session.delete(student)
        if user:
            db.session.delete(user)
        db.session.commit()
        flash(f'Student {std_id} deleted successfully.', 'success')
    else:
        flash('Student not found.', 'error')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete-faculty/<fac_id>', methods=['POST'])
@login_required
def delete_faculty(fac_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    faculty = Faculty.query.get(fac_id)
    if faculty:
        user = User.query.get(faculty.user_id)
        db.session.delete(faculty)
        if user:
            db.session.delete(user)
        db.session.commit()
        flash(f'Faculty {fac_id} deleted successfully.', 'success')
    else:
        flash('Faculty not found.', 'error')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/create-section-offering', methods=['POST'])
@login_required
def create_section_offering():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    ccode = request.form.get('course_code', '').strip()
    snum = request.form.get('section_number', '').strip()
    credits = float(request.form.get('credits', 3.0))
    cap = int(request.form.get('capacity', 30))
    sched = request.form.get('schedule', '').strip()
    room = request.form.get('room', '').strip()
    fac_id = request.form.get('faculty_id') or None
    is_lab = 'is_lab' in request.form
    linked_id = request.form.get('linked_section_id') or None
    
    dept_str = request.form.get('dedicated_departments', '').strip()
    dedicated = [x.strip() for x in dept_str.split('-') if x.strip()] if dept_str else []
    
    pre_str = request.form.get('prerequisites', '').strip()
    prereqs = [x.strip() for x in pre_str.split(',') if x.strip()] if pre_str else []
    
    valid_buildings = ['AB1', 'AB2', 'AB3', 'FUB', 'Main']
    building_ok = False
    for b in valid_buildings:
        if room.startswith(b):
            building_ok = True
            break
            
    if not building_ok:
        flash(f"Invalid Building in Room field. Room must start with AB1, AB2, AB3, FUB, or Main.", 'error')
        return redirect(url_for('admin_dashboard'))
        
    sec_id = f"{ccode}-{snum}-{get_next_semester()}"
    
    existing = SectionOffering.query.filter_by(id=sec_id).first()
    if existing:
        flash(f"Section offering '{sec_id}' already exists in the system.", 'error')
        return redirect(url_for('admin_dashboard'))
        
    comp_cred_req = 0
    try:
        comp_cred_req = int(request.form.get('completed_credit_requirement', '0'))
    except ValueError:
        pass

    sec = SectionOffering(
        id=sec_id,
        course_code=ccode,
        course_title=ccode + (" Lab" if is_lab else " Course"),
        section_number=snum,
        credits=credits,
        capacity=cap,
        schedule=sched,
        room=room,
        faculty_id=fac_id,
        is_lab=is_lab,
        linked_section_id=linked_id,
        semester_id=get_next_semester(),
        completed_credit_requirement=comp_cred_req
    )
    sec.dedicated_departments = dedicated
    sec.prerequisites = prereqs
    
    db.session.add(sec)
    
    if linked_id:
        linked_sec = SectionOffering.query.get(linked_id)
        if linked_sec:
            linked_sec.linked_section_id = sec_id
            
    db.session.commit()
    
    flash('Section offering created successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/create-pre-course', methods=['POST'])
@login_required
def create_pre_course():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    code = request.form.get('code', '').strip()
    title = request.form.get('title', '').strip()
    credits = float(request.form.get('credits', 3.0))
    dept = request.form.get('department_id')
    pre_str = request.form.get('prerequisites', '').strip()
    prereqs = [x.strip() for x in pre_str.split(',') if x.strip()] if pre_str else []
    
    existing = PreAdvisingCourse.query.filter_by(id=code).first()
    if existing:
        flash(f"Pre-advising course '{code}' already exists.", 'error')
        return redirect(url_for('admin_dashboard'))
        
    pc = PreAdvisingCourse(
        id=code,
        code=code,
        title=title,
        credits=credits,
        department_id=dept
    )
    pc.prerequisites = prereqs
    db.session.add(pc)
    db.session.commit()
    
    flash('Course catalog added successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/create-window', methods=['POST'])
@login_required
def create_window():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    win_type = request.form.get('type')
    label = request.form.get('label')
    c_min = float(request.form.get('credit_min'))
    c_max = float(request.form.get('credit_max'))
    start = request.form.get('start_date_time')
    end = request.form.get('end_date_time')
    
    win_id = f"win-{win_type}-{int(datetime.utcnow().timestamp())}"
    win = AdvisingWindow(
        id=win_id,
        type=win_type,
        label=label,
        credit_min=c_min,
        credit_max=c_max,
        start_date_time=start,
        end_date_time=end,
        semester_id=get_current_semester()
    )
    db.session.add(win)
    db.session.commit()
    
    flash('Advising timeline slot added successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

def import_excel_schedule(file_source):
    import openpyxl
    wb = openpyxl.load_workbook(file_source)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        header_row = next(sheet.iter_rows(max_row=1), None)
        if not header_row:
            continue
            
        headers = [str(cell.value).strip() for cell in header_row]
        required_cols = ['ID', 'Course Code', 'Section', 'Date & Time', 'Credit', 'Seat Capacity', 'Dedicated Department', 'Room', 'Pre-Requisite', 'Linked Course ID', 'Completed Credit Requirement']
        
        col_map = {}
        for name in required_cols:
            if name in headers:
                col_map[name] = headers.index(name)
            else:
                for idx, h in enumerate(headers):
                    if h.lower().replace(' ', '').replace('-', '') == name.lower().replace(' ', '').replace('-', ''):
                        col_map[name] = idx
                        break
                        
        fac_col = None
        for idx, h in enumerate(headers):
            if h.lower().replace(' ', '').replace('_', '').replace('-', '') in ['faculty', 'facultyid', 'assignedfaculty']:
                fac_col = idx
                break
                
        if 'Course Code' not in col_map or 'Section' not in col_map:
            continue
            
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        row_mapping = {}
        semester_id = get_next_semester()
        
        # Pass 1: Add/Update Courses and Sections
        for row in rows:
            if not row or all(v is None for v in row):
                continue
                
            row_id = row[col_map['ID']] if 'ID' in col_map else None
            ccode = str(row[col_map['Course Code']]).strip() if row[col_map['Course Code']] is not None else ''
            if not ccode or ccode.lower() == 'none':
                continue
                
            section_val = row[col_map['Section']]
            if section_val is None:
                continue
            snum = f"{int(section_val):02d}"
            sec_id = f"{ccode}-{snum}-{semester_id}"
            if row_id is not None:
                row_mapping[row_id] = sec_id
                
            credits_val = float(row[col_map['Credit']]) if ('Credit' in col_map and row[col_map['Credit']] is not None) else 3.0
            
            dept_str = str(row[col_map['Dedicated Department']]).strip() if ('Dedicated Department' in col_map and row[col_map['Dedicated Department']]) else ''
            depts = [d.strip() for d in dept_str.split(',') if d.strip()]
            first_dept = depts[0] if depts else 'CSE'
            
            prereq_str = str(row[col_map['Pre-Requisite']]).strip() if ('Pre-Requisite' in col_map and row[col_map['Pre-Requisite']]) else ''
            prereqs = [p.strip() for p in prereq_str.split(',') if p.strip() and p.strip().lower() != 'none']
            
            comp_cred_req = 0
            if 'Completed Credit Requirement' in col_map and row[col_map['Completed Credit Requirement']] is not None:
                try:
                    comp_cred_req = int(row[col_map['Completed Credit Requirement']])
                except ValueError:
                    comp_cred_req = 0

            # Parse Faculty Assignment from Excel if present
            fac_id = None
            if fac_col is not None and row[fac_col] is not None:
                fac_val = str(row[fac_col]).strip()
                if fac_val and fac_val.lower() != 'none':
                    faculty = Faculty.query.get(fac_val)
                    if not faculty:
                        faculty = Faculty.query.filter(Faculty.id.ilike(fac_val)).first()
                    if not faculty:
                        faculty = Faculty.query.filter(Faculty.name.ilike(fac_val)).first()
                    if faculty:
                        fac_id = faculty.id
            
            # Course Catalog sync
            course = PreAdvisingCourse.query.get(ccode)
            if not course:
                course = PreAdvisingCourse(
                    id=ccode,
                    code=ccode,
                    title=ccode,
                    credits=credits_val,
                    department_id=first_dept,
                    completed_credit_requirement=comp_cred_req
                )
                course.prerequisites = prereqs
                db.session.add(course)
            else:
                course.credits = credits_val
                course.department_id = first_dept
                course.prerequisites = prereqs
                course.completed_credit_requirement = comp_cred_req
                
            # SectionOffering sync
            sched = str(row[col_map['Date & Time']]).strip() if ('Date & Time' in col_map and row[col_map['Date & Time']]) else 'TBA'
            room = str(row[col_map['Room']]).strip() if ('Room' in col_map and row[col_map['Room']]) else 'TBA'
            
            cap_val = str(row[col_map['Seat Capacity']]).strip() if ('Seat Capacity' in col_map and row[col_map['Seat Capacity']] is not None) else '30'
            capacity = int(cap_val.split('/')[-1]) if '/' in cap_val else int(cap_val)
            
            is_lab = 'Lab' in ccode
            
            sec = SectionOffering.query.get(sec_id)
            if not sec:
                sec = SectionOffering(
                    id=sec_id,
                    course_code=ccode,
                    course_title=ccode,
                    section_number=snum,
                    credits=credits_val,
                    schedule=sched,
                    room=room,
                    capacity=capacity,
                    is_lab=is_lab,
                    semester_id=semester_id,
                    completed_credit_requirement=comp_cred_req
                )
                sec.dedicated_departments = depts
                sec.prerequisites = prereqs
                db.session.add(sec)
            else:
                sec.credits = credits_val
                sec.schedule = sched
                sec.room = room
                sec.capacity = capacity
                sec.dedicated_departments = depts
                sec.prerequisites = prereqs
                sec.completed_credit_requirement = comp_cred_req
            
            if fac_id:
                # Validate schedule conflict for this faculty
                conflict = False
                assigned_secs = SectionOffering.query.filter_by(faculty_id=fac_id, semester_id=semester_id).all()
                for a in assigned_secs:
                    if a.id != sec_id and schedules_conflict(a.schedule, sched):
                        conflict = True
                        break
                if not conflict:
                    sec.faculty_id = fac_id
                else:
                    print(f"Warning: Faculty {fac_id} has conflict with section {sec_id} in sheet schedule.")
                
        db.session.commit()
        
        # Pass 2: Setup linked Lab/Theory relations
        if 'Linked Course ID' in col_map:
            for row in rows:
                if not row or all(v is None for v in row):
                    continue
                row_id = row[col_map['ID']] if 'ID' in col_map else None
                linked_val = row[col_map['Linked Course ID']]
                if row_id is None or linked_val is None or str(linked_val).strip().lower() == 'none':
                    continue
                    
                try:
                    linked_row_id = int(linked_val)
                except ValueError:
                    continue
                    
                sec_id = row_mapping.get(row_id)
                linked_sec_id = row_mapping.get(linked_row_id)
                if sec_id and linked_sec_id:
                    sec = SectionOffering.query.get(sec_id)
                    if sec:
                        sec.linked_section_id = linked_sec_id
            db.session.commit()

def import_excel_students(file_source):
    import openpyxl
    import re
    wb = openpyxl.load_workbook(file_source)
    sheet = wb.active
    header_row = next(sheet.iter_rows(max_row=1), None)
    if not header_row:
        raise ValueError("The excel file is empty or has no header row.")
        
    headers = [str(cell.value).strip() for cell in header_row]
    
    required_cols = [
        'Student ID', 'Name', 'Student Email', 'Phone Number', 'Department',
        'Completed Credit', 'Remaining Credit', 'CGPA', 'Present Address',
        'Permanent Address', 'Completed Courses and Grades', 'Current Courses',
        'Current Course Credit', 'Next Semester Courses', 'Next Semester Course Credit',
        'Profile Picture'
    ]
    
    col_map = {}
    for name in required_cols:
        if name in headers:
            col_map[name] = headers.index(name)
        else:
            for idx, h in enumerate(headers):
                if h.lower().replace(' ', '').replace('_', '').replace('-', '') == name.lower().replace(' ', '').replace('_', '').replace('-', ''):
                    col_map[name] = idx
                    break
                    
    if 'Student ID' not in col_map or 'Name' not in col_map or 'Student Email' not in col_map:
        raise ValueError("Headers must contain at least 'Student ID', 'Name', and 'Student Email'.")
        
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    imported_count = 0
    
    for row in rows:
        if not row or all(v is None for v in row):
            continue
            
        std_id = str(row[col_map['Student ID']]).strip()
        name = str(row[col_map['Name']]).strip()
        email = str(row[col_map['Student Email']]).strip()
        
        if not std_id or not name or not email:
            continue
            
        phone = str(row[col_map['Phone Number']]).strip() if ('Phone Number' in col_map and row[col_map['Phone Number']] is not None) else None
        dept = str(row[col_map['Department']]).strip() if ('Department' in col_map and row[col_map['Department']] is not None) else 'CSE'
        
        try:
            credits = float(row[col_map['Completed Credit']]) if ('Completed Credit' in col_map and row[col_map['Completed Credit']] is not None) else 0.0
        except ValueError:
            credits = 0.0
            
        try:
            rem_credits = float(row[col_map['Remaining Credit']]) if ('Remaining Credit' in col_map and row[col_map['Remaining Credit']] is not None) else 140.0
        except ValueError:
            rem_credits = 140.0
            
        try:
            cgpa = float(row[col_map['CGPA']]) if ('CGPA' in col_map and row[col_map['CGPA']] is not None) else 0.0
        except ValueError:
            cgpa = 0.0
            
        present_address = str(row[col_map['Present Address']]).strip() if ('Present Address' in col_map and row[col_map['Present Address']] is not None) else None
        permanent_address = str(row[col_map['Permanent Address']]).strip() if ('Permanent Address' in col_map and row[col_map['Permanent Address']] is not None) else None
        
        comp_courses_grades = str(row[col_map['Completed Courses and Grades']]).strip() if ('Completed Courses and Grades' in col_map and row[col_map['Completed Courses and Grades']] is not None) else None
        curr_courses = str(row[col_map['Current Courses']]).strip() if ('Current Courses' in col_map and row[col_map['Current Courses']] is not None) else None
        
        try:
            curr_credit = float(row[col_map['Current Course Credit']]) if ('Current Course Credit' in col_map and row[col_map['Current Course Credit']] is not None) else 0.0
        except ValueError:
            curr_credit = 0.0
            
        next_courses = str(row[col_map['Next Semester Courses']]).strip() if ('Next Semester Courses' in col_map and row[col_map['Next Semester Courses']] is not None) else None
        
        try:
            next_credit = float(row[col_map['Next Semester Course Credit']]) if ('Next Semester Course Credit' in col_map and row[col_map['Next Semester Course Credit']] is not None) else 0.0
        except ValueError:
            next_credit = 0.0
            
        prof_pic = str(row[col_map['Profile Picture']]).strip() if ('Profile Picture' in col_map and row[col_map['Profile Picture']] is not None) else None
        if prof_pic and prof_pic.lower() == 'none':
            prof_pic = None
            
        user = User.query.filter_by(email=email).first()
        if not user:
            user = User(
                id='usr-' + std_id,
                email=email,
                password_hash=generate_password_hash('password123'),
                role='student',
                is_active=True,
                is_activated=False
            )
            db.session.add(user)
            db.session.flush()
            
        student = Student.query.get(std_id)
        if not student:
            student = Student(
                id=std_id,
                user_id=user.id,
                name=name,
                department_id=dept,
                completed_credits=credits,
                remaining_credits=rem_credits,
                cgpa=cgpa,
                phone_number=phone,
                present_address=present_address,
                permanent_address=permanent_address,
                completed_courses_and_grades=comp_courses_grades,
                current_courses=curr_courses,
                current_course_credit=curr_credit,
                next_semester_courses=next_courses,
                next_semester_course_credit=next_credit,
                profile_pic=prof_pic,
                outstanding_balance=0,
                financial_cleared=True,
                about=''
            )
            db.session.add(student)
        else:
            student.name = name
            student.department_id = dept
            student.completed_credits = credits
            student.remaining_credits = rem_credits
            student.cgpa = cgpa
            student.phone_number = phone
            student.present_address = present_address
            student.permanent_address = permanent_address
            student.completed_courses_and_grades = comp_courses_grades
            student.current_courses = curr_courses
            student.current_course_credit = curr_credit
            student.next_semester_courses = next_courses
            student.next_semester_course_credit = next_credit
            if prof_pic:
                student.profile_pic = prof_pic
                
        if comp_courses_grades and comp_courses_grades.lower() != 'none':
            parts = re.split(r'[,;\n\r]+', comp_courses_grades)
            for part in parts:
                part = part.strip()
                if not part: continue
                match = re.search(r'([A-Za-z0-9\s]+)[:\s]+([A-Za-z+-]+)', part)
                if match:
                    ccode = match.group(1).strip().replace(' ', '')
                    gletter = match.group(2).strip()
                    points_map = {
                        'A+': 4.0, 'A': 4.0, 'A-': 3.7,
                        'B+': 3.3, 'B': 3.0, 'B-': 2.7,
                        'C+': 2.3, 'C': 2.0, 'C-': 1.7,
                        'D+': 1.3, 'D': 1.0, 'F': 0.0
                    }
                    gpoint = points_map.get(gletter.upper(), 0.0)
                    grade_id = f"GRD-{std_id}-{ccode}"
                    
                    existing_grade = Grade.query.get(grade_id)
                    if not existing_grade:
                        new_grade = Grade(
                            id=grade_id,
                            student_id=std_id,
                            section_id=ccode,
                            grade_letter=gletter.upper(),
                            grade_point=gpoint,
                            semester_id='completed'
                        )
                        db.session.add(new_grade)
                    else:
                        existing_grade.grade_letter = gletter.upper()
                        existing_grade.grade_point = gpoint
                        
        if curr_courses and curr_courses.lower() != 'none':
            ccodes = [c.strip() for c in re.split(r'[,;\s]+', curr_courses) if c.strip()]
            for cc in ccodes:
                if not cc: continue
                sec = SectionOffering.query.filter_by(course_code=cc).first()
                if sec:
                    reg_id = f"REG-{std_id}-{sec.id}"
                    existing_reg = Registration.query.get(reg_id)
                    if not existing_reg:
                        new_reg = Registration(
                            id=reg_id,
                            student_id=std_id,
                            section_id=sec.id,
                            semester_id=sec.semester_id,
                            status='registered'
                        )
                        db.session.add(new_reg)
                        sec.enrolled_count = min(sec.capacity, sec.enrolled_count + 1)
                        
        imported_count += 1
        
    db.session.commit()
    return imported_count

@app.route('/admin/upload-students', methods=['POST'])
@login_required
def upload_students():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    file = request.files.get('student_file')
    if not file or file.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=students')
        
    if not file.filename.endswith('.xlsx'):
        flash('Invalid file format. Only Excel files (.xlsx) are allowed.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=students')
        
    try:
        count = import_excel_students(file)
        flash(f'{count} students imported/updated successfully!', 'success')
    except Exception as e:
        flash(f'Error importing students: {str(e)}', 'error')
        
    return redirect(url_for('admin_dashboard') + '?tab=students')

@app.route('/admin/upload-schedule', methods=['POST'])
@login_required
def upload_schedule():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=course-management')
        
    if not file.filename.endswith('.xlsx'):
        flash('Invalid file format. Only Excel files (.xlsx) are allowed.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=course-management')
        
    try:
        import_excel_schedule(file)
        flash('Course schedule imported successfully!', 'success')
    except Exception as e:
        flash(f'Error importing schedule: {str(e)}', 'error')
        
    return redirect(url_for('admin_dashboard') + '?tab=course-management')

@app.route('/admin/assign-faculty-section', methods=['POST'])
@login_required
def assign_faculty_section():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    fac_id = request.form.get('faculty_id')
    sec_id = request.form.get('section_id')
    
    faculty = Faculty.query.get(fac_id)
    section = SectionOffering.query.get(sec_id)
    
    if not faculty:
        flash('Faculty member not found.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=faculty')
        
    if not section:
        flash('Course section offering not found.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=faculty')
        
    # Check for schedule conflicts
    next_sem = get_next_semester()
    assigned = SectionOffering.query.filter_by(faculty_id=fac_id, semester_id=next_sem).all()
    for a in assigned:
        if a.id != sec_id and schedules_conflict(a.schedule, section.schedule):
            flash(f"Conflict detected! This section ({section.course_code} Sec {section.section_number}: {section.schedule}) conflicts with {a.course_code} Sec {a.section_number} ({a.schedule}) which is already assigned to {faculty.name}.", "error")
            return redirect(url_for('admin_dashboard') + '?tab=faculty')
            
    # Assign
    section.faculty_id = fac_id
    db.session.commit()
    flash(f"Successfully assigned {section.course_code} Sec {section.section_number} ({section.schedule}) to {faculty.name}.", "success")
    return redirect(url_for('admin_dashboard') + '?tab=faculty')

@app.route('/admin/assign-faculty-sections-bulk', methods=['POST'])
@login_required
def assign_faculty_sections_bulk():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    faculty_id = request.form.get('faculty_id')
    section_ids = request.form.getlist('section_ids')
    
    faculty = Faculty.query.get(faculty_id)
    if not faculty:
        flash("Faculty member not found.", "error")
        return redirect(url_for('admin_dashboard') + "?tab=faculty")
        
    if not section_ids:
        flash("No sections selected for assignment.", "error")
        return redirect(url_for('admin_dashboard') + "?tab=faculty")
        
    success_count = 0
    conflict_count = 0
    semester_id = get_next_semester()
    
    assigned_secs = SectionOffering.query.filter(
        SectionOffering.faculty_id == faculty_id,
        SectionOffering.semester_id == semester_id,
        ~SectionOffering.id.in_(section_ids)
    ).all()
    
    to_assign = []
    current_schedule_list = [(s.id, s.schedule, f"{s.course_code}-Sec{s.section_number}") for s in assigned_secs]
    
    for sec_id in section_ids:
        sec = SectionOffering.query.get(sec_id)
        if not sec:
            continue
            
        has_conflict = False
        for aid, asched, acode in current_schedule_list:
            if schedules_conflict(asched, sec.schedule):
                has_conflict = True
                flash(f"Conflict: Section {sec.course_code}-Sec{sec.section_number} ({sec.schedule}) conflicts with already assigned {acode} ({asched}).", "error")
                break
        
        if not has_conflict:
            for tid, tsched, tcode in to_assign:
                if schedules_conflict(tsched, sec.schedule):
                    has_conflict = True
                    flash(f"Conflict: Section {sec.course_code}-Sec{sec.section_number} ({sec.schedule}) conflicts with selected {tcode} ({tsched}).", "error")
                    break
                    
        if not has_conflict:
            to_assign.append((sec.id, sec.schedule, f"{sec.course_code}-Sec{sec.section_number}"))
            sec.faculty_id = faculty_id
            success_count += 1
        else:
            conflict_count += 1
            
    db.session.commit()
    
    if success_count > 0:
        flash(f"Successfully assigned {success_count} section(s) to {faculty.name}.", "success")
    if conflict_count > 0:
        flash(f"{conflict_count} section(s) could not be assigned due to scheduling conflicts.", "error")
        
    return redirect(url_for('admin_dashboard') + "?tab=faculty")

@app.route('/admin/unassign-faculty-section/<sec_id>', methods=['POST'])
@login_required
def unassign_faculty_section(sec_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    section = SectionOffering.query.get(sec_id)
    if not section:
        flash('Course section offering not found.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=faculty')
        
    fac_id = section.faculty_id
    faculty = Faculty.query.get(fac_id) if fac_id else None
    fac_name = faculty.name if faculty else "Faculty"
    
    section.faculty_id = None
    db.session.commit()
    flash(f"Successfully unassigned {section.course_code} Sec {section.section_number} from {fac_name}.", "success")
    return redirect(url_for('admin_dashboard') + '?tab=faculty')

@app.route('/admin/delete-window/<win_id>', methods=['POST'])
@login_required
def delete_window(win_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    AdvisingWindow.query.filter_by(id=win_id).delete()
    db.session.commit()
    flash('Advising window timeline deleted.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/add-pre-course', methods=['POST'])
@login_required
def add_pre_course():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    code = request.form.get('code', '').strip().upper()
    try:
        credits = float(request.form.get('credits', '3.0'))
    except ValueError:
        flash('Invalid credits value.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=pre-advising')
        
    if not code:
        flash('Course Code is required.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=pre-advising')
        
    existing = PreAdvisingCourse.query.filter_by(code=code).first()
    if existing:
        flash(f'Course {code} already exists in the pre-advising catalog.', 'error')
        return redirect(url_for('admin_dashboard') + '?tab=pre-advising')
        
    import re
    match = re.match(r'^([A-Za-z]+)', code)
    dept = match.group(1) if match else 'GEN'
    
    comp_cred_req = 0
    try:
        comp_cred_req = int(request.form.get('completed_credit_requirement', '0'))
    except ValueError:
        pass

    course = PreAdvisingCourse(
        id=code,
        code=code,
        title=code,
        credits=credits,
        department_id=dept,
        _prerequisites='[]',
        completed_credit_requirement=comp_cred_req
    )
    db.session.add(course)
    db.session.commit()
    flash(f'Course {code} ({credits} Credits) added successfully with credit requirement {comp_cred_req}.', 'success')
    return redirect(url_for('admin_dashboard') + '?tab=pre-advising')

@app.route('/admin/delete-pre-course/<course_id>', methods=['POST'])
@login_required
def delete_pre_course(course_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    course = PreAdvisingCourse.query.get(course_id)
    if course:
        db.session.delete(course)
        db.session.commit()
        flash(f'Course {course_id} removed from the pre-advising catalog.', 'success')
    else:
        flash('Course not found.', 'error')
        
    return redirect(url_for('admin_dashboard') + '?tab=pre-advising')

@app.route('/admin/delete-section-offering/<sec_id>', methods=['POST'])
@login_required
def delete_section_offering(sec_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    SectionOffering.query.filter_by(id=sec_id).delete()
    db.session.commit()
    flash('Section offering deleted successfully.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/post-announcement', methods=['POST'])
@login_required
def post_announcement():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    title = request.form.get('title')
    target = request.form.get('target_role')
    content = request.form.get('content')
    
    ann = Announcement(
        id=f"ann-{int(datetime.utcnow().timestamp())}",
        title=title,
        content=content,
        created_by="EWU Registrar Office",
        created_at=datetime.utcnow().strftime('%Y-%m-%d'),
        target_role=target
    )
    db.session.add(ann)
    db.session.commit()
    
    flash('Notice announcement published!', 'success')
    return redirect(url_for('admin_dashboard'))

def recalculate_student_stats(student_id):
    student = Student.query.get(student_id)
    if not student:
        return
    grades = Grade.query.filter_by(student_id=student_id).all()
    if not grades:
        student.cgpa = 0.0
        student.completed_credits = 0.0
        db.session.commit()
        return
    
    total_credits = 0.0
    weighted_points = 0.0
    for g in grades:
        course = PreAdvisingCourse.query.filter_by(code=g.section_id).first()
        cr = course.credits if course else 3.0
        total_credits += cr
        weighted_points += g.grade_point * cr
        
    student.cgpa = weighted_points / total_credits if total_credits > 0 else 0.0
    passed_credits = sum((PreAdvisingCourse.query.filter_by(code=g.section_id).first().credits if PreAdvisingCourse.query.filter_by(code=g.section_id).first() else 3.0) for g in grades if g.grade_point > 0.0)
    student.completed_credits = passed_credits
    db.session.commit()

@app.route('/admin/add-grade', methods=['POST'])
@login_required
def add_grade():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    student_id = request.form.get('student_id', '').strip()
    course_code = request.form.get('course_code', '').strip().upper()
    grade_letter = request.form.get('grade_letter', '').strip().upper()
    semester_id = request.form.get('semester_id', '').strip()
    
    grade_points_map = {
        'A': 4.0, 'A-': 3.7, 'B+': 3.3, 'B': 3.0, 'B-': 2.7,
        'C+': 2.3, 'C': 2.0, 'C-': 1.7, 'D+': 1.3, 'D': 1.0, 'F': 0.0
    }
    
    if grade_letter not in grade_points_map:
        flash(f"Invalid grade letter '{grade_letter}'. Choose from A, A-, B+, B, B-, C+, C, C-, D+, D, F.", 'error')
        return redirect(url_for('admin_dashboard'))
        
    student = Student.query.get(student_id)
    if not student:
        flash(f"Student ID '{student_id}' does not exist.", 'error')
        return redirect(url_for('admin_dashboard'))
        
    grade_point = grade_points_map[grade_letter]
    
    grade = Grade.query.filter_by(student_id=student_id, section_id=course_code).first()
    if not grade:
        grade = Grade(
            id=f"grade-{student_id}-{course_code}",
            student_id=student_id,
            section_id=course_code,
            grade_letter=grade_letter,
            grade_point=grade_point,
            semester_id=semester_id
        )
        db.session.add(grade)
    else:
        grade.grade_letter = grade_letter
        grade.grade_point = grade_point
        grade.semester_id = semester_id
        
    db.session.commit()
    
    recalculate_student_stats(student_id)
    
    flash(f"Grade '{grade_letter}' saved for {student_id} in {course_code} successfully! Academic records updated.", 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete-grade/<grade_id>', methods=['POST'])
@login_required
def delete_grade(grade_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
        
    grade = Grade.query.get(grade_id)
    if grade:
        student_id = grade.student_id
        db.session.delete(grade)
        db.session.commit()
        recalculate_student_stats(student_id)
        flash("Grade entry deleted and stats recalculated.", "success")
    else:
        flash("Grade not found.", "error")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/student-details/<std_id>')
@login_required
def admin_student_details(std_id):
    if current_user.role != 'admin':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
        
    student = Student.query.get(std_id)
    if not student:
        return jsonify({'status': 'error', 'message': 'Student not found.'}), 404
        
    user = User.query.get(student.user_id)
    email = user.email if user else 'N/A'
    
    requests = AdvisingRequest.query.filter_by(student_id=student.id).all()
    requests_list = [{
        'course_id': r.course_id,
        'type': r.type,
        'status': r.status,
        'created_at': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''
    } for r in requests]
    
    regs = Registration.query.filter_by(student_id=student.id).all()
    regs_list = []
    for r in regs:
        sec = SectionOffering.query.get(r.section_id)
        if sec:
            regs_list.append({
                'course_code': sec.course_code,
                'section_number': sec.section_number,
                'semester_id': r.semester_id,
                'schedule': sec.schedule,
                'room': sec.room
            })
            
    grades = Grade.query.filter_by(student_id=student.id).all()
    grades_list = [{
        'id': g.id,
        'course_code': g.section_id,
        'grade_letter': g.grade_letter,
        'grade_point': g.grade_point,
        'semester_id': g.semester_id
    } for g in grades]
    
    advisor = Faculty.query.get(student.advisor_id) if student.advisor_id else None
    advisor_name = advisor.name if advisor else 'Not Assigned'
    
    ledger = LedgerEntry.query.filter_by(student_id=student.id).all()
    ledger_list = [{
        'description': l.description,
        'amount': l.amount,
        'status': l.status,
        'date': l.date
    } for l in ledger]
    
    return jsonify({
        'status': 'success',
        'id': student.id,
        'name': student.name,
        'email': email,
        'department_id': student.department_id,
        'completed_credits': student.completed_credits,
        'cgpa': student.cgpa,
        'outstanding_balance': student.outstanding_balance,
        'financial_cleared': student.financial_cleared,
        'advising_status': student.advising_status,
        'about': student.about or '',
        'profile_pic': student.profile_pic or '',
        'advisor_name': advisor_name,
        'registrations': regs_list,
        'grades': grades_list,
        'ledger': ledger_list,
        
        # New profile fields
        'phone_number': student.phone_number or '',
        'remaining_credits': student.remaining_credits,
        'present_address': student.present_address or '',
        'permanent_address': student.permanent_address or '',
        'completed_courses_and_grades': student.completed_courses_and_grades or '',
        'current_courses': student.current_courses or '',
        'current_course_credit': student.current_course_credit,
        'next_semester_courses': student.next_semester_courses or '',
        'next_semester_course_credit': student.next_semester_course_credit
    })

@app.route('/admin/faculty-details/<fac_id>')
@login_required
def admin_faculty_details(fac_id):
    if current_user.role != 'admin':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
        
    faculty = Faculty.query.get(fac_id)
    if not faculty:
        return jsonify({'status': 'error', 'message': 'Faculty not found.'}), 404
        
    user = User.query.get(faculty.user_id)
    email = user.email if user else 'N/A'
    
    sections = SectionOffering.query.filter_by(faculty_id=faculty.id).all()
    sections_list = [{
        'course_code': s.course_code,
        'section_number': s.section_number,
        'semester_id': s.semester_id,
        'schedule': s.schedule,
        'room': s.room,
        'capacity': s.capacity,
        'enrolled_count': s.enrolled_count
    } for s in sections]
    
    advisees = Student.query.filter_by(advisor_id=faculty.id).all()
    advisees_list = [{
        'id': std.id,
        'name': std.name,
        'department_id': std.department_id,
        'cgpa': std.cgpa
    } for std in advisees]
    
    return jsonify({
        'status': 'success',
        'id': faculty.id,
        'name': faculty.name,
        'email': email,
        'department_id': faculty.department_id,
        'about': faculty.about or '',
        'profile_pic': faculty.profile_pic or '',
        'sections': sections_list,
        'advisees': advisees_list
    })

if __name__ == '__main__':
    app.run(debug=True, port=3001)
